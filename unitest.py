
# -*- coding:utf-8 -*-
from http.client import responses
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from time import sleep, perf_counter, perf_counter
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from pathlib import Path
from datetime import date, datetime
from lxml.etree import ParserError
import pandas as pd
import requests
import random

# user_agent = UserAgent()
# headers = {'User-Agent': user_agent.random}

def get_dividend(url,pair_id):
    """股息URL測試"""
    if "?cid=" in url:
        f_index = url.find("?cid=")
        l_index = url[f_index:]
        url = url[:f_index]+"-dividends"+l_index
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
    else:
        url = url+"-dividends"
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
    finaldf = pd.DataFrame()
    df = pd.read_html(str(xml))
    for i in df:
        if "除息日" in i.columns:
            finaldf = i
    dividend_list = list(dataframe_to_rows(finaldf, index=False, header=False))
    if dividend_list:
        last_timestamp = xml.find_all("td",class_="left first")[-1].attrs["data-value"]
        while True:
            tr_list, last_timestamp = get_more_dividend(pair_id=pair_id,last_timestamp=last_timestamp)
            if tr_list:
                for i in tr_list:
                    all_ele = i.contents
                    date1 = all_ele[1].text
                    date1_value = all_ele[3].text
                    date2 = all_ele[7].text
                    date2_value = all_ele[9].text
                    append_list = [ date1,date1_value,"",date2,date2_value]
                    dividend_list.append(append_list)
                sleep(0.5)
            else:
                break
    return dividend_list

def get_more_dividend(pair_id,last_timestamp):
    url = "https://hk.investing.com/equities/MoreDividendsHistory"
    stock_data = { "pairID": pair_id, "last_timestamp": last_timestamp }
    headers = { "User-Agent": user_agent.random,
                "x-requested-with": "XMLHttpRequest" }
    res = requests.post(url, headers=headers, data=stock_data)
    res.encoding = "UTF-8"
    history_row = res.json()["historyRows"]
    xml = BeautifulSoup(history_row, "html.parser")
    try:
        tr_list= xml.find_all("tr")
        last_timestamp = tr_list[-1].find("td",class_="left first").attrs["data-value"]
    except:
        tr_list= ""
        last_timestamp = ""
    return tr_list,last_timestamp

# response = get_more_dividend(pair_id="103731",last_timestamp=1468368000)
# print(response)

# print(get_dividend(url="https://hk.investing.com/equities/formosa-plasti",pair_id="103008"))

# profitloss = 0


if __name__ == "__main__":
    wb = load_workbook("(8150)南茂科技2022-06-07.xlsx")
    ws = wb["股息整理"]
    print(ws["J2"].value)




