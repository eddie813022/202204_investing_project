
# -*- coding:utf-8 -*-
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from time import sleep, perf_counter, perf_counter
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from pathlib import Path
from datetime import date, datetime
from lxml.etree import ParserError
import pandas as pd
import requests
import random

user_agent = UserAgent()
headers = {'User-Agent': user_agent.random}


def ransleep():
    sleep(random.randint(1, 2))


def autowidth(ws):
    lks = []
    for i in range(1, ws.max_column+1):
        lk = 1
        for j in range(1, ws.max_row + 1):
            sz = ws.cell(row=j, column=i).value
            if isinstance(sz, str):
                lk1 = len(sz.encode('utf-8'))
            else:
                lk1 = len(str(sz))
            if lk < lk1:
                lk = lk1
        lks.append(lk)
    for i in range(1, ws.max_column + 1):
        k = get_column_letter(i)
        ws.column_dimensions[k].width = lks[i-1]+2


def styled_cells(data, ws, key1_list, key2_list=[""], key3_list=[""]):
    font1 = Font(color="0066CC")
    font2 = Font(color="006030")
    font3 = Font(bold=True)
    for i in data:
        if i in key1_list:
            value = Cell(ws, column="A", row=1, value=i)
            value.font = font1
        elif i in key2_list:
            value = Cell(ws, column="A", row=1, value=i)
            value.font = font2
        else:
            if i in key3_list:
                value = Cell(ws, column="A", row=1, value=i)
                value.font = font3
            else:
                value = Cell(ws, column="A", row=1, value=i)
        yield i


def ws1_write_column(ws):
    ws1_column_list = ["收盤價", "全日波幅", "營業額", "開市", "52週波幅", "EPS", "成交量", "市值",
                       "股息(紅利)", "平均成交量(3個月)", "市盈率", "Beta", "1年生跌率", "已發行股票",
                       "下一個財務報告公布", "公司名稱", "產業", "板塊", "股票類型"]
    for i in range(19):
        ws.cell(row=i+1, column=1, value=ws1_column_list[0])
        ws1_column_list.pop(0)


def ws2_write_column(ws):
    ws2_column_list = ["年度", "最高", "最低"]
    for i in range(3):
        ws.cell(row=i+1, column=1, value=ws2_column_list[0])
        ws2_column_list.pop(0)


def ws3to4_write_column(ws1, ws2):
    ws3_column_list_1 = ["毛利率TTM", "經營利潤率TTM", "淨利率TTM", "投資回報率TTM"]
    ws3_column_list_2 = ["速動比率MRQ", "流動比率MRQ", "長期負債股權比MRQ", "總負債股權比MRQ"]
    ws3_column_list_3 = ["每股現金流TTM", "每股收入TTM", "業務現金流"]
    for i in range(4):
        ws1.cell(row=1, column=i+1, value=ws3_column_list_1[0])
        ws2.cell(row=1, column=i+1, value=ws3_column_list_1[0])
        ws3_column_list_1.pop(0)
    for i in range(4):
        ws1.cell(row=3, column=i+1, value=ws3_column_list_2[0])
        ws2.cell(row=3, column=i+1, value=ws3_column_list_2[0])
        ws3_column_list_2.pop(0)
    for i in range(3):
        ws1.cell(row=5, column=i+1, value=ws3_column_list_3[0])
        ws2.cell(row=5, column=i+1, value=ws3_column_list_3[0])
        ws3_column_list_3.pop(0)


def ws11_write_column(ws):
    ws11_column_list = ["名稱", "公司", "產業"]
    ws.cell(row=1, column=1, value=ws11_column_list[0])
    ws.cell(row=1, column=2, value=ws11_column_list[1])
    ws.cell(row=1, column=3, value=ws11_column_list[2])


def ws12_write_column(ws):
    ws12_column_list = ["除息日", "股息", "種類", "付息日", "孳息率"]
    ws.cell(row=1, column=1, value=ws12_column_list[0])
    ws.cell(row=1, column=2, value=ws12_column_list[1])
    ws.cell(row=1, column=3, value=ws12_column_list[2])
    ws.cell(row=1, column=4, value=ws12_column_list[3])
    ws.cell(row=1, column=5, value=ws12_column_list[4])


def ws13_write_column(ws):
    ws13_column_list = ["公佈日期", "截止", "EPS/預測", "營業額/預測"]
    ws.cell(row=1, column=1, value=ws13_column_list[0])
    ws.cell(row=1, column=2, value=ws13_column_list[1])
    ws.cell(row=1, column=3, value=ws13_column_list[2])
    ws.cell(row=1, column=4, value=ws13_column_list[3])


"""抓取國家名稱&代碼清單"""


def get_country_dict():
    basic_url = "https://hk.investing.com/stock-screener/?sp=country::39|sector::a|industry::a|equityType::a%3Ceq_market_cap;1"
    res = requests.get(basic_url, headers=headers)
    res.encoding = "UTF-8"
    country_xml = BeautifulSoup(res.text, "lxml").find(
        id="countriesUL").find_all("li")
    country_dict = {}
    for i in country_xml:
        country_dict[i.text.replace("\n", "").replace(
            " ", "")] = i["data-value"]
    return country_dict


"""抓取市場名稱&代碼清單"""


def get_exchang_dict(country_dict):
    exchange_category_dict = {}
    for j, k in country_dict.items():
        exchange_url = f"https://hk.investing.com/stock-screener/?sp=country::{k}|sector::a|industry::a|equityType::a%3Ceq_market_cap;1"
        exchange_res = requests.get(exchange_url, headers=headers)
        exchange_res.encoding = "UTF-8"
        exchange_xml = BeautifulSoup(exchange_res.text, "html.parser").find(
            id="exchangesUL").find_all("li")
        for i in exchange_xml:
            exchange_name = i.text
            exchange_code = i["data-value"]
            exchange_category_dict[exchange_name] = exchange_code
        return exchange_category_dict


"""抓取股票清單"""


def get_stock_list(**kwargs):
    start = perf_counter()
    wb = Workbook()
    ws = wb.active
    ws.title = "個股清單"
    country = kwargs["ccode"]
    exchange = kwargs["ecode"]
    category = kwargs["gcode"]
    savename = kwargs["save"]
    progress = kwargs["progress"]
    stock_data = {"country[]": country, "exchange[]": exchange,
                  "sector": category, "order[col]": "eq_market_cap", "order[dir]": "d"}
    headers = {"User-Agent": user_agent.random, "x-requested-with": "XMLHttpRequest",
               "referer": f"https://hk.investing.com/stock-screener/?sp=country::{country}|sector::{category}|industry::a|equityType::a|exchange::{exchange}%3Ceq_market_cap;1"}
    basic_url = "https://hk.investing.com/stock-screener/Service/SearchStocks"
    res = requests.post(basic_url, headers=headers, data=stock_data)
    res.encoding = "UTF-8"
    res = res.json()
    stock_totalcount = res["totalCount"]
    stock_list = []
    if int(stock_totalcount):
        stock_count = int(stock_totalcount)
        divisor = 50                        # 除數
        quotient = stock_count / divisor    # 商數
        remainder = stock_count % divisor   # 餘數
        sinteger = stock_count // divisor  # 整除數
        if remainder == 0:
            for i in res["hits"]:
                stock_list.append(
                    (i["viewData"]["name"], "https://hk.investing.com"+i["viewData"]["link"], i["pair_ID"]))
        else:
            sinteger += 1
            update_progres = int(100 / sinteger)
            for page in range(sinteger):
                stock_data = {"country[]": country, "exchange[]": exchange, "sector": category,
                              "pn": page+1, "order[col]": "eq_market_cap", "order[dir]": "d"}
                res = requests.post(
                    basic_url, headers=headers, data=stock_data)
                res.encoding = "UTF-8"
                res = res.json()
                for i in res["hits"]:
                    stock_list.append(
                        (i["viewData"]["name"], "https://hk.investing.com"+i["viewData"]["link"], i["pair_ID"]))
                progress["value"] += update_progres
                ransleep()
    else:
        progress["value"] = 100
        wb.close()
    if stock_list:
        for i in range(len(stock_list)):
            ws.cell(row=i+1, column=1, value=stock_list[0][0])
            ws.cell(row=i+1, column=2, value=stock_list[0][1])
            ws.cell(row=i+1, column=3, value=stock_list[0][2])
            stock_list.pop(0)
        wb.save(savename)
        progress["value"] = 100
        print(f"Cost: {perf_counter() - start}")
    return stock_totalcount


"""讀取複選框設定"""


def get_checkboxconfig():
    tempfile = Path.cwd() / "temp.xlsx"
    wb = load_workbook(tempfile)
    ws = wb.active
    f1_1config = [ws.cell(row=i, column=2).value for i in range(2, 35)]  # 33
    f1_2config = [ws.cell(row=i, column=2).value for i in range(35, 69)]  # 34
    f1_3config = [ws.cell(row=i, column=2).value for i in range(69, 93)]  # 24
    f2_1config = [ws.cell(row=i, column=2).value for i in range(93, 140)]  # 47
    f2_2config = [
        ws.cell(row=i, column=2).value for i in range(140, 185)]  # 45
    f2_3config = [
        ws.cell(row=i, column=2).value for i in range(185, 227)]  # 42
    f3config = [ws.cell(row=i, column=2).value for i in range(227, 253)]  # 26
    wb.close()
    return f1_1config, f1_2config, f1_3config, f2_1config, f2_2config, f2_3config, f3config


"""讀取日期設定"""


def get_dateconfig():
    f1_1list = ["總收入", "收入", "其他收入合計", "稅收成本合計", "毛利", "經營開支總額", "銷售/一般/管理費用合計",
                "研發", "折舊/攤銷", "利息開支(收入)-營運淨額", "例外開支(收入)", "其他運營開支總額", "營業收入",
                "利息收入（開支）- 非營運淨額", "出售資產收入（虧損）", "其他，淨額", "稅前淨收益", "備付所得稅",
                "稅後淨收益", "少數股東權益", "附屬公司權益", "美國公認會計準則調整", "計算特殊項目前的淨收益", "特殊項目合計",
                "淨收入", "淨收入調整總額", "扣除特殊項目的普通收入", "稀釋調整", "稀釋后淨收入", "稀釋后加權平均股",
                "稀釋后扣除特殊項目的每股盈利", "每股股利 – 普通股首次發行", "稀釋后每股標準盈利"]
    f1_2list = ["總收入", "保費收入合計", "投資收益淨額", "變現收益（虧損）", "其他收入合計", "經營開支總額", "虧損、福利和修訂合計",
                "购置成本攤銷", "銷售/一般/管理費用合計", "折舊/攤銷", "利息開支（收入）- 營運淨額", "例外開支（收入）", "其他運營開支總額",
                "營業收入", "利息收入（開支）- 非營運淨額", "出售資產收入（虧損）", "其他，淨額", "稅前淨收益", "備付所得稅",
                "稅後淨收益", "少數股東權益", "附屬公司權益", "美國公認會計準則調整", "計算特殊項目前的淨收益", "特殊項目合計",
                "淨收入", "淨收入調整總額", "扣除特殊項目的普通收入", "稀釋調整", "稀釋后淨收入", "稀釋后加權平均股",
                "稀釋后扣除特殊項目的每股盈利", "每股股利 – 普通股首次發行", "稀釋后每股標準盈利"]
    f1_3list = ["利息收益淨額", "銀行利息收入", "利息開支總額", "風險準備金", "扣除風險準備金後淨利息收入", "銀行非利息收入",
                "銀行非利息開支", "稅前淨收益", "備付所得稅", "稅後淨收益", "少數股東權益", "附屬公司權益", "美國公認會計準則調整",
                "計算特殊項目前的淨收益", "特殊項目合計", "淨收入", "淨收入調整總額", "扣除特殊項目的普通收入", "稀釋調整",
                "稀釋后淨收入", "稀釋后加權平均股", "稀釋后扣除特殊項目的每股盈利", "每股股利 – 普通股首次發行", "稀釋后每股標準盈利"]
    f2_1_list = ["流動資產合計", "現金和短期投資", "現金", "現金和現金等價物", "短期投資", "淨應收款合計", "淨交易應收款合計",
                 "庫存合計", "預付費用", "其他流動資產合計", "總資產", "物業/廠房/設備淨總額", "物業/廠房/設備總額",
                 "累計折舊合計", "商譽淨額", "無形資產淨額", "長期投資", "長期應收票據", "其他長期資產合計", "其他資產合計",
                 "總流動負債", "應付賬款", "應付/應計", "應計費用", "應付票據/短期債務", "長期負債當前應收部分/資本租賃",
                 "其他流動負債合計", "總負債", "長期債務合計", "長期債務", "資本租賃債務", "遞延所得稅", "少數股東權益",
                 "其他負債合計", "總權益", "可贖回優先股合計", "不可贖回優先股淨額", "普通股合計", "附加資本", "保留盈餘(累計虧損)",
                 "普通庫存股", "員工持股計劃債務擔保", "未實現收益（虧損）", "其他權益合計", "負債及股東權益總計",
                 "已發行普通股合計", "已發行優先股合計"]
    f2_2_list = ["流動資產合計", "總資產", "現金", "現金和現金等價物", "淨應收款合計", "預付費用", "物業/廠房/設備淨總額", "物業/廠房/設備總額", "累計折舊合計",
                 "商譽淨額", "無形資產淨額", "長期投資", "應收保險", "長期應收票據", "其他長期資產合計", "遞延保單獲得成本", "其他資產合計", "總流動負債",
                 "總負債", "應付賬款", "應付/應計", "應計費用", "保單負債", "應付票據/短期債務", "長期負債當前應收部分/資本租賃", "其他流動負債合計",
                 "長期債務合計", "長期債務", "資本租賃債務", "遞延所得稅", "少數股東權益", "其他負債合計", "總權益", "可贖回優先股合計", "不可贖回優先股淨額",
                 "普通股合計", "附加資本", "保留盈餘(累計虧損)", "普通庫存股", "員工持股計劃債務擔保", "未實現收益（虧損）", "其他權益合計",
                 "負債及股東權益總計", "已發行普通股合計", "已發行優先股合計"]
    f2_3_list = ["流動資產合計", "總資產", "銀行應付現金和欠款", "其他盈利資產合計", "淨貸款", "物業/廠房/設備淨總額", "物業/廠房/設備總額", "累計折舊合計",
                 "商譽淨額", "無形資產淨額", "長期投資", "其他長期資產合計", "其他資產合計", "總流動負債", "總負債", "應付賬款", "應付/應計", "應計費用",
                 "存款總額", "其他付息負債合計", "短期借貸總額", "長期負債當前應收部分/資本租賃", "其他流動負債合計", "長期債務合計", "長期債務", "資本租賃債務",
                 "遞延所得稅", "少數股東權益", "其他負債合計", "總權益", "可贖回優先股合計", "不可贖回優先股淨額", "普通股合計", "附加資本", "保留盈餘(累計虧損)",
                 "普通庫存股", "員工持股計劃債務擔保", "未實現收益（虧損）", "其他權益合計", "負債及股東權益總計", "已發行普通股合計", "已發行優先股合計"]
    f3_list = ["淨收益/起點", "來自經營活動的現金", "折舊/遞耗", "攤銷", "遞延稅", "非現金項目", "現金收入", "現金支出", "現金稅金支出",
               "現金利息支出", "營運資金變動", "來自投資活動的現金", "資本支出", "其他投資現金流項目合計", "來自融資活動的現金",
               "融資現金流項目", "發放現金紅利合計", "股票發行（贖回）淨額", "債務發行（贖回）淨額", "外匯影響", "現金變動淨額",
               "期初現金結餘", "期末現金結餘", "自由現金流", "自由現金流增長", "自由現金流收益率"]
    tempfile = Path.cwd() / "temp.xlsx"
    nowday = date.today().strftime("%Y/%m/%d")
    if tempfile.is_file():
        wb = load_workbook(tempfile)
        ws = wb.active
        date_ = ws.cell(row=1, column=2).value
        singlepath = ws.cell(row=253, column=2).value
        bulkpath = ws.cell(row=254, column=2).value
        wb.close()
    else:
        wb = Workbook()
        ws = wb.active
        date_ = nowday
        ws.cell(row=1, column=1, value="from_date")
        ws.cell(row=1, column=2, value=date_)
        for i in range(33):  # 損益1
            ws.cell(row=i+2, column=1).value = f1_1list[0]
            ws.cell(row=i+2, column=2).value = "True"
            f1_1list.pop(0)
        for i in range(34):  # 損益2
            ws.cell(row=i+35, column=1).value = f1_2list[0]
            ws.cell(row=i+35, column=2).value = "True"
            f1_2list.pop(0)
        for i in range(24):  # 損益3
            ws.cell(row=i+69, column=1).value = f1_3list[0]
            ws.cell(row=i+69, column=2).value = "True"
            f1_3list.pop(0)
        for i in range(47):  # 資產1
            ws.cell(row=i+93, column=1).value = f2_1_list[0]
            ws.cell(row=i+93, column=2).value = "True"
            f2_1_list.pop(0)
        for i in range(45):  # 資產2
            ws.cell(row=i+140, column=1).value = f2_2_list[0]
            ws.cell(row=i+140, column=2).value = "True"
            f2_2_list.pop(0)
        for i in range(42):  # 資產3
            ws.cell(row=i+185, column=1).value = f2_3_list[0]
            ws.cell(row=i+185, column=2).value = "True"
            f2_3_list.pop(0)
        for i in range(26):  # 現金
            ws.cell(row=i+227, column=1).value = f3_list[0]
            ws.cell(row=i+227, column=2).value = "True"
            f3_list.pop(0)
        ws.cell(row=253, column=1, value="singledownload")
        ws.cell(row=253, column=2, value="C:/")
        ws.cell(row=254, column=1, value="bulkdownload")
        ws.cell(row=254, column=2, value="C:/")
        wb.save(tempfile)
    year = int(date_[0:4])
    month = int(date_[5:date_.rfind("/")])
    day = int(date_[date_.rfind("/")+1:])
    return year, month, day


"""寫入日期設定"""


def write_dateconfig(st_date):
    tempfile = Path.cwd() / "temp.xlsx"
    st_date = str(st_date).replace("-", "/")
    wb = load_workbook(tempfile)
    ws = wb.active
    ws.cell(row=1, column=2, value=st_date)
    wb.save(tempfile)


"""讀取下載路徑設定"""


def get_pathconfig():
    tempfile = Path.cwd() / "temp.xlsx"
    wb = load_workbook(tempfile)
    ws = wb.active
    single = ws.cell(row=253, column=2).value
    bulk = ws.cell(row=254, column=2).value
    return single, bulk


"""寫入下載路徑設定"""


def write_pathconfig(type, download_path):
    tempfile = Path.cwd() / "temp.xlsx"
    wb = load_workbook(tempfile)
    ws = wb.active
    if type == "single":
        ws.cell(row=253, column=2, value=download_path)
    elif type == "bulk":
        ws.cell(row=254, column=2, value=download_path)
    wb.save(tempfile)


"""讀取個股清單"""


def read_stock_list(stockpath):
    wb = load_workbook(stockpath)
    ws = wb.active
    try:
        for i in range(ws.max_row):
            if ws.cell(row=i+1, column=4).value != "finished":
                last_index = ws.cell(row=i+1, column=4).row
                wb.close()
                return last_index
        last_index = 0
        wb.close()
        return last_index
    except:
        wb.close()
        last_index = 0
        return last_index


"""抓取綜觀"""
# https://hk.investing.com/equities/william-demant


def get_stock_inventory(url, ws):
    res = requests.get(url, headers=headers)
    res.encoding = "UTF-8"
    xml = BeautifulSoup(res.text, "lxml")
    h1parser = xml.find("h1").get_text()
    # TODO網站確認有抓到，是否為try-except時出問題導致抓到空值
    stockpair = h1parser[h1parser.find("("):h1parser.find(")")+1]
    ws1_column_list = []
    try:
        prev_close = xml.find(
            attrs={"data-test": "prevClose"}).find("span").text
    except:
        prev_close = ""
    ws1_column_list.append(prev_close)
    try:
        day_range = xml.find(attrs={"data-test": "dailyRange"}).find_all("span")[
            0].text + " - " + xml.find(attrs={"data-test": "dailyRange"}).find_all("span")[4].text
    except:
        day_range = ""
    ws1_column_list.append(day_range)
    try:
        revenue = xml.find(attrs={"data-test": "revenue"}).find("span").text
    except:
        revenue = ""
    ws1_column_list.append(revenue)
    try:
        prev_open = xml.find(attrs={"data-test": "open"}).find("span").text
    except:
        prev_open = ""
    ws1_column_list.append(prev_open)
    try:
        weak52_range = xml.find(attrs={"data-test": "weekRange"}).find_all("span")[
            0].text + " - " + xml.find(attrs={"data-test": "dailyRange"}).find_all("span")[4].text
    except:
        weak52_range = ""
    ws1_column_list.append(weak52_range)
    try:
        eps = xml.find(attrs={"data-test": "eps"}).find("span").text
    except:
        eps = ""
    ws1_column_list.append(eps)
    try:
        volume = xml.find(attrs={"data-test": "volume"}).find("span").text
    except:
        volume = ""
    ws1_column_list.append(volume)
    try:
        market_cup = xml.find(
            attrs={"data-test": "marketCap"}).find("span").text
    except:
        market_cup = ""
    ws1_column_list.append(market_cup)
    try:
        dividend = xml.find(attrs={"data-test": "dividend"}).find("span").text
        if dividend == "(":
            dividend = "N/A(N/A)"
    except:
        dividend = ""
    ws1_column_list.append(dividend)
    try:
        avgvolume = xml.find(
            attrs={"data-test": "avgVolume"}).find("span").text
    except:
        avgvolume = ""
    ws1_column_list.append(avgvolume)
    try:
        ratio = xml.find(attrs={"data-test": "ratio"}).find("span").text
    except:
        ratio = ""
    ws1_column_list.append(ratio)
    try:
        beta = xml.find(attrs={"data-test": "beta"}).find("span").text
    except:
        beta = "-"
    ws1_column_list.append(beta)
    try:
        oneyear_return = xml.find(
            attrs={"data-test": "oneYearReturn"}).find("span").text
    except:
        oneyear_return = ""
    ws1_column_list.append(oneyear_return)
    try:
        shareout = xml.find(
            attrs={"data-test": "sharesOutstanding"}).find("span").text
    except:
        shareout = ""
    ws1_column_list.append(shareout)
    try:
        nextdate = xml.find(
            attrs={"data-test": "nextEarningDate"}).find("a").text
    except:
        nextdate = ""
    ws1_column_list.append(nextdate)
    for i in range(len(ws1_column_list)):
        ws.cell(row=i+1, column=2, value=ws1_column_list[0])
        ws1_column_list.pop(0)
    return stockpair


"""抓取簡介"""
# "https://hk.investing.com/equities/william-demant-profile"


def get_stock_profile(url,ws):
    if "?cid=" in url:
        f_index = url.find("?cid=")
        l_index = url[f_index:]
        url = url[:f_index]+"-company-profile"+l_index
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
    else:
        url = url+"-company-profile"
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
    try:
        stock_name = xml.select_one("h1").text.replace(" ", "")
    except:
        index = url.rfind("/")
        stock_name = url[index:]
    try:
        industry = xml.find(class_="companyProfileHeader").find("a").text
    except:
        industry = ""
    try:
        category = xml.find(
            class_="companyProfileHeader").contents[3].find("a").text
    except:
        category = ""
    try:
        stocktype = xml.find(
            class_="companyProfileHeader").contents[7].find("p").text
    except:
        category = ""
    ws.cell(row=16, column=2, value=stock_name)
    ws.cell(row=17, column=2, value=industry)
    ws.cell(row=18, column=2, value=category)
    ws.cell(row=19, column=2, value=stocktype)


"""抓取歷史數據"""
# "https://hk.investing.com/instruments/HistoricalDataAjax"


def get_stock_history(pair_id, start_date, end_date, ws):
    if start_date == end_date:
        yearindex = str(int(start_date[:4])-4)
        start_date = yearindex+start_date[4:]
    url = "https://hk.investing.com/instruments/HistoricalDataAjax"
    headers = {"User-Agent": user_agent.random,
               "x-requested-with": "XMLHttpRequest"}
    stock_data = {"curr_id": pair_id, "action": "historical_data", "interval_sec": "Monthly",
                  "sort_ord": "DESC", "st_date": start_date, "end_date": end_date}
    res = requests.post(url, headers=headers, data=stock_data)
    res.encoding = "UTF-8"
    xml = BeautifulSoup(res.text, "lxml").find(id="curr_table")
    df = pd.read_html(str(xml))[0]
    from_year = start_date[0:4]
    # from_month = int(start_date[start_date.find("/")+1:start_date.rfind("/")])
    end_year = end_date[0:4]
    # end_month = int(end_date[end_date.find("/")+1:end_date.rfind("/")])
    count_year = (int(end_year) - int(from_year)) + 1
    end_year_n = int(end_year) + 1
    df["日期"] = df["日期"].str[0:4]
    df_column = df["日期"] == str(int(end_year_n)-1)
    dfmax = df[df_column]["高"].max(axis=0, skipna=True)
    dfmin = df[df_column]["低"].min(axis=0, skipna=True)
    year_list = []
    max_list = []
    min_list = []
    if count_year >= 1:
        for i in range(count_year):
            try:
                df_column = df["日期"] == str(int(end_year_n-(i+1)))
                yearname = str(int(end_year_n-(i+1)))
                dfmax = df[df_column]["高"].max(axis=0, skipna=True)
                dfmin = df[df_column]["低"].min(axis=0, skipna=True)
            except:
                yearname = "-"
                dfmax = "-"
                dfmin = "-"
            year_list.append(yearname)
            max_list.append(dfmax)
            min_list.append(dfmin)
        for i in range(len(year_list)):
            ws.cell(row=1, column=i+2, value=year_list[0])
            year_list.pop(0)
        for i in range(len(max_list)):
            ws.cell(row=2, column=i+2, value=max_list[0])
            max_list.pop(0)
        for i in range(len(min_list)):
            ws.cell(row=3, column=i+2, value=min_list[0])
            min_list.pop(0)


"""抓取財務摘要"""
# https://hk.investing.com/instruments/Financials/changesummaryreporttypeajax


def get_stock_financials(pair_id, ws1, ws2):
    url = "https://hk.investing.com/instruments/Financials/changesummaryreporttypeajax"
    headers = {"User-Agent": user_agent.random,
               "x-requested-with": "XMLHttpRequest"}
    stock_data_annual = {"pid": pair_id, "action": "change_report_type",
                         "financial_id": pair_id, "ratios_id": pair_id, "period_type": "Interim"}
    stock_data_interim = {"pid": pair_id, "action": "change_report_type",
                          "financial_id": pair_id, "ratios_id": pair_id, "period_type": "Annual"}
    res = requests.post(url, headers=headers, data=stock_data_annual)
    sleep(0.5)
    res2 = requests.post(url, headers=headers, data=stock_data_interim)
    res.encoding = "UTF-8"
    res2.encoding = "UTF-8"
    xml = BeautifulSoup(res.text, "lxml")
    xml2 = BeautifulSoup(res2.text, "lxml")
    xmlspan = xml.find_all("span")
    xml2span = xml2.find_all("span")
    df1 = pd.read_html(str(xml))[0]
    df2 = pd.read_html(str(xml))[1]
    df3 = pd.read_html(str(xml))[2]
    df4 = pd.read_html(str(xml2))[0]
    df5 = pd.read_html(str(xml2))[1]
    df6 = pd.read_html(str(xml2))[2]
    avalue1 = xmlspan[2].text
    avalue2 = xmlspan[5].text
    avalue3 = xmlspan[8].text
    avalue4 = xmlspan[8].text
    avalue5 = xmlspan[20].text
    avalue6 = xmlspan[23].text
    avalue7 = xmlspan[26].text
    avalue8 = xmlspan[29].text
    avalue9 = xmlspan[38].text
    avalue10 = xmlspan[41].text
    avalue11 = xmlspan[44].text
    ivalue1 = xml2span[2].text
    ivalue2 = xml2span[5].text
    ivalue3 = xml2span[8].text
    ivalue4 = xml2span[8].text
    ivalue5 = xml2span[20].text
    ivalue6 = xml2span[23].text
    ivalue7 = xml2span[26].text
    ivalue8 = xml2span[29].text
    ivalue9 = xml2span[38].text
    ivalue10 = xml2span[41].text
    ivalue11 = xml2span[44].text
    avalue1_list = [avalue1, avalue2, avalue3, avalue4]
    avalue2_list = [avalue5, avalue6, avalue7, avalue8]
    avalue3_list = [avalue9, avalue10, avalue11]
    ivalue1_list = [ivalue1, ivalue2, ivalue3, ivalue4]
    ivalue2_list = [ivalue5, ivalue6, ivalue7, ivalue8]
    ivalue3_list = [ivalue9, ivalue10, ivalue11]
    for i in range(4):
        ws1.cell(row=2, column=i+1, value=avalue1_list[0])
        ws2.cell(row=2, column=i+1, value=ivalue1_list[0])
        avalue1_list.pop(0)
        ivalue1_list.pop(0)
    for i in range(4):
        ws1.cell(row=4, column=i+1, value=avalue2_list[0])
        ws2.cell(row=4, column=i+1, value=ivalue2_list[0])
        avalue2_list.pop(0)
        ivalue2_list.pop(0)
    for i in range(3):
        ws1.cell(row=6, column=i+1, value=avalue3_list[0])
        ws2.cell(row=6, column=i+1, value=ivalue3_list[0])
        avalue3_list.pop(0)
        ivalue3_list.pop(0)
    ws1.append([""])
    ws2.append([""])
    for r in dataframe_to_rows(df1, index=False, header=True):
        if len(r) > 1:
            ws1.append(r)
    ws1.append([""])
    for r in dataframe_to_rows(df2, index=False, header=True):
        if len(r) > 1:
            ws1.append(r)
    ws1.append([""])
    df3_list = list(dataframe_to_rows(df3, index=False, header=True))
    df3_list.pop(0)
    data3 = list(df3.head(0))
    data3.pop(0)
    data3.insert(0, ("結束日期：", ""))
    endofdate3 = [i[0] for i in data3]
    df3_list.insert(0, endofdate3)
    for r in df3_list:
        if len(r) > 1:
            ws1.append(r)
    for r in dataframe_to_rows(df4, index=False, header=True):
        if len(r) > 1:
            ws2.append(r)
    ws2.append([""])
    for r in dataframe_to_rows(df5, index=False, header=True):
        if len(r) > 1:
            ws2.append(r)
    ws2.append([""])
    df6_list = list(dataframe_to_rows(df6, index=False, header=True))
    df6_list.pop(0)
    data6 = list(df6.head(0))
    data6.pop(0)
    data6.insert(0, ("結束日期：", ""))
    endofdate6 = [i[0] for i in data6]
    df6_list.insert(0, endofdate6)
    for r in df6_list:
        if len(r) > 1:
            ws2.append(r)


"""抓取損益表"""
# https://hk.investing.com/instruments/Financials/changereporttypeajax


def get_profitandloss(pair_id, choose1_list, choose2_list, choose3_list, ws1, ws2):
    profit1_list = choose1_list
    profit2_list = choose2_list
    profit3_list = choose3_list
    general_list = ["收入", "其他收入合計", "銷售/一般/管理費用合計", "研發",
                    "折舊/攤銷", "利息開支（收入）- 營運淨額", "例外開支（收入）", "其他運營開支總額"]
    finance1_list = ["保費收入合計", "投資收益淨額", "變現收益（虧損）", "其他收入合計", "虧損、福利和修訂合計", "购置成本攤銷",
                     "銷售/一般/管理費用合計", "折舊/攤銷", "利息開支（收入）- 營運淨額", "例外開支（收入）", "其他運營開支總額"]
    finance2_list = ["銀行利息收入", "利息開支總額"]
    url = "https://hk.investing.com/instruments/Financials/changereporttypeajax"
    headers = {"User-Agent": user_agent.random,
               "x-requested-with": "XMLHttpRequest"}
    stock_data_annual = {"pair_ID": pair_id, "action": "change_report_type",
                         "report_type": "INC", "period_type": "Interim"}
    stock_data_interim = {"pair_ID": pair_id, "action": "change_report_type",
                          "report_type": "INC", "period_type": "Annual"}
    res = requests.post(url, headers=headers, data=stock_data_annual)
    sleep(0.5)
    res2 = requests.post(url, headers=headers, data=stock_data_interim)
    res.encoding = "UTF-8"
    res2.encoding = "UTF-8"
    xml = BeautifulSoup(res.text, "lxml")
    xml2 = BeautifulSoup(res2.text, "lxml")
    savetype = "其它"
    try:
        df1 = pd.read_html(str(xml))[0]
    except ParserError as f:
        df1 = pd.DataFrame()
        profitloss = 0
        error = "季報表，"+str(f)
    except Exception as e:
        df1 = pd.DataFrame()
        profitloss = 0
        error = "季報表，"+e
    try:
        df2 = pd.read_html(str(xml2))[0]
    except ParserError as f:
        df1 = pd.DataFrame()
        profitloss = 0
        error = "年報表，"+str(f)
    except Exception as e:
        df2 = pd.DataFrame()
        profitloss = 0
        error = "年報表，"+e
    if not df1.empty and not df2.empty:
        if len(df1.index) == 36:
            bold_list = ["總收入", "毛利", "經營開支總額", "營業收入", "稅前淨收益", "稅後淨收益",
                         "計算特殊項目前的淨收益", "淨收入", "扣除特殊項目的普通收入", "稀釋后扣除特殊項目的每股盈利"]
            frame1_to_rows = list(dataframe_to_rows(
                df1, index=False, header=False))
            end_date = frame1_to_rows[-1]
            clean_end_date = []
            for date in end_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
                    clean_end_date.append(date_string)
                if str(date) == "nan":
                    clean_end_date.append("")
            clean_end_date[0] = "結束日期："
            frame1_to_rows.pop(1)
            frame1_to_rows.pop(6)
            frame1_to_rows.pop(-1)
            frame1_to_rows.insert(0, clean_end_date)
            frame2_to_rows = list(dataframe_to_rows(
                df2, index=False, header=False))
            end2_date = frame2_to_rows[-1]
            clean_end2_date = []
            for date in end2_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
                    clean_end2_date.append(date_string)
                if str(date) == "nan":
                    clean_end2_date.append("")
            clean_end2_date[0] = "結束日期："
            frame2_to_rows.pop(1)
            frame2_to_rows.pop(6)
            frame2_to_rows.pop(-1)
            frame2_to_rows.insert(0, clean_end2_date)
            for i in frame1_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            for i in frame2_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            if len(profit1_list) != len(frame1_to_rows):
                profit1_list.insert(0, "1")
            final1_rows = []
            final2_rows = []
            for i in profit1_list:
                if i:
                    final1_rows.append(frame1_to_rows[0])
                frame1_to_rows.pop(0)
            for i in profit1_list:
                if i:
                    final2_rows.append(frame2_to_rows[0])
                frame2_to_rows.pop(0)
            for i in final1_rows:
                ws1.append(styled_cells(
                    data=i, key1_list=general_list, key3_list=bold_list, ws=ws1))
            for i in final2_rows:
                ws2.append(styled_cells(
                    data=i, key1_list=general_list, key3_list=bold_list, ws=ws2))
            profitloss = 0
            error = 0
            savetype = "一般"
        elif len(df1.index) == 37:
            bold_list = ["總收入", "經營開支總額", "營業收入", "稅前淨收益", "稅後淨收益",
                         "計算特殊項目前的淨收益", "淨收入", "扣除特殊項目的普通收入", "稀釋后扣除特殊項目的每股盈利"]
            frame1_to_rows = list(dataframe_to_rows(
                df1, index=False, header=False))
            end_date = frame1_to_rows[-1]
            clean_end_date = []
            for date in end_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
                    clean_end_date.append(date_string)
                if str(date) == "nan":
                    clean_end_date.append("")
            clean_end_date[0] = "結束日期："
            frame1_to_rows.pop(1)
            frame1_to_rows.pop(6)
            frame1_to_rows.pop(-1)
            frame1_to_rows.insert(0, clean_end_date)
            frame2_to_rows = list(dataframe_to_rows(
                df2, index=False, header=False))
            end2_date = frame2_to_rows[-1]
            clean_end2_date = []
            for date in end2_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
                    clean_end2_date.append(date_string)
                if str(date) == "nan":
                    clean_end2_date.append("")
            clean_end2_date[0] = "結束日期："
            frame2_to_rows.pop(1)
            frame2_to_rows.pop(6)
            frame2_to_rows.pop(-1)
            frame2_to_rows.insert(0, clean_end2_date)
            for i in frame1_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            for i in frame2_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            if len(profit2_list) != len(frame1_to_rows):
                profit2_list.insert(0, "1")
            final1_rows = []
            final2_rows = []
            for i in profit2_list:
                if i:
                    final1_rows.append(frame1_to_rows[0])
                frame1_to_rows.pop(0)
            for i in profit2_list:
                if i:
                    final2_rows.append(frame2_to_rows[0])
                frame2_to_rows.pop(0)
            for i in final1_rows:
                ws1.append(styled_cells(
                    data=i, key1_list=finance1_list, key3_list=bold_list, ws=ws1))
            for i in final2_rows:
                ws2.append(styled_cells(
                    data=i, key1_list=finance1_list, key3_list=bold_list, ws=ws2))
            profitloss = 0
            error = 0
            savetype = "保險"
        elif len(df1.index) == 26:
            bold_list = ["利息收益淨額", "扣除風險準備金後淨利息收入", "稅前淨收益", "稅後淨收益",
                         "計算特殊項目前的淨收益", "淨收入", "扣除特殊項目的普通收入", "稀釋后扣除特殊項目的每股盈利"]
            frame1_to_rows = list(dataframe_to_rows(
                df1, index=False, header=False))
            end_date = frame1_to_rows[-1]
            clean_end_date = []
            for date in end_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
                    clean_end_date.append(date_string)
                if str(date) == "nan":
                    clean_end_date.append("")
            clean_end_date[0] = "結束日期："
            frame1_to_rows.pop(1)
            frame1_to_rows.pop(-1)
            frame1_to_rows.insert(0, clean_end_date)
            frame2_to_rows = list(dataframe_to_rows(
                df2, index=False, header=False))
            end2_date = frame2_to_rows[-1]
            clean_end2_date = []
            for date in end2_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
                    clean_end2_date.append(date_string)
                if str(date) == "nan":
                    clean_end2_date.append("")
            clean_end2_date[0] = "結束日期："
            frame2_to_rows.pop(1)
            frame2_to_rows.pop(-1)
            frame2_to_rows.insert(0, clean_end2_date)
            for i in frame1_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            for i in frame2_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            if len(profit3_list) != len(frame1_to_rows):
                profit3_list.insert(0, "1")
            final1_rows = []
            final2_rows = []
            for i in profit3_list:
                if i:
                    final1_rows.append(frame1_to_rows[0])
                frame1_to_rows.pop(0)
            for i in profit3_list:
                if i:
                    final2_rows.append(frame2_to_rows[0])
                frame2_to_rows.pop(0)
            for i in final1_rows:
                ws1.append(styled_cells(
                    data=i, key1_list=finance2_list, key3_list=bold_list, ws=ws1))
            for i in final2_rows:
                ws2.append(styled_cells(
                    data=i, key1_list=finance2_list, key3_list=bold_list, ws=ws2))
            profitloss = 0
            error = 0
            savetype = "銀行"
        else:
            error = "df行數" + str(len(df1.index))
            profitloss = 1
            # try:
            #     frame1_to_rows = list(dataframe_to_rows(
            #         df1, index=False, header=False))
            #     end_date = frame1_to_rows[-1]
            #     clean_end_date = []
            #     for date in end_date:
            #         if str(date) != "nan":
            #             date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
            #             clean_end_date.append(date_string)
            #         if str(date) == "nan":
            #             clean_end_date.append("")
            #     clean_end_date[0] = "結束日期："
            #     frame1_to_rows.pop(1)
            #     frame1_to_rows.pop(-1)
            #     frame1_to_rows.insert(0, clean_end_date)
            #     frame2_to_rows = list(dataframe_to_rows(
            #         df2, index=False, header=False))
            #     end_date = frame2_to_rows[-1]
            #     clean_end2_date = []
            #     for date in end2_date:
            #         if str(date) != "nan":
            #             date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
            #             clean_end2_date.append(date_string)
            #         if str(date) == "nan":
            #             clean_end2_date.append("")
            #     clean_end2_date[0] = "結束日期："
            #     frame2_to_rows.pop(1)
            #     frame2_to_rows.pop(-1)
            #     frame2_to_rows.insert(0, clean_end2_date)
            #     for i in frame1_to_rows:
            #         for index, value in enumerate(i):
            #             if i[index] == "-":
            #                 i[index] = ""
            #     for i in frame2_to_rows:
            #         for index, value in enumerate(i):
            #             if i[index] == "-":
            #                 i[index] = ""
            #     for i in final1_rows:
            #         ws1.append(styled_cells(
            #             data=i, key1_list=finance2_list, key3_list=bold_list, ws=ws1))
            #     for i in final2_rows:
            #         ws2.append(styled_cells(
            #             data=i, key1_list=finance2_list, key3_list=bold_list, ws=ws2))
            #     profitloss = 1
            #     error = 0
            # except Exception as e:
            #     frame1_to_rows = []
            #     frame2_to_rows = []
            #     profitloss = 1
            #     error = e
            #     savetype = "其它"
    return int(profitloss), error, savetype


"""抓取資產負債表"""
# https://hk.investing.com/instruments/Financials/changereporttypeajax


def get_stock_balance(pair_id, choose1_list, choose2_list, choose3_list, ws1, ws2):
    balance1_list = choose1_list
    balance2_list = choose2_list
    balance3_list = choose3_list
    general1_1_list = ["現金和短期投資", "淨應收款合計", "庫存合計", "預付費用", "其他流動資產合計",
                       "物業/廠房/設備淨總額", "商譽淨額", "無形資產淨額", "長期投資",
                       "長期應收票據", "其他長期資產合計", "其他資產合計", "應付賬款",
                       "應付/應計", "應計費用", "應付票據/短期債務", "長期負債當前應收部分/資本租賃",
                       "其他流動負債合計", "長期債務合計", "遞延所得稅", "少數股東權益", "其他負債合計",
                       "可贖回優先股合計", "不可贖回優先股淨額", "普通股合計", "附加資本", "保留盈餘(累計虧損)",
                       "普通庫存股", "員工持股計劃債務擔保", "未實現收益（虧損）", "其他權益合計"]
    general1_2_list = ["現金", "現金和現金等價物", "短期投資", "淨交易應收款合計",
                       "物業/廠房/設備總額", "累計折舊合計", "長期債務", "資本租賃債務"]
    finance1_1_list = ["現金", "現金和現金等價物", "淨應收款合計", "預付費用", "物業/廠房/設備淨總額", "商譽淨額", "無形資產淨額", "長期投資",
                       "應收保險", "長期應收票據", "其他長期資產合計", "遞延保單獲得成本", "其他資產合計", "應付賬款", "應付/應計", "應計費用",
                       "保單負債", "應付票據/短期債務", "長期負債當前應收部分/資本租賃", "其他流動負債合計", "長期債務合計", "遞延所得稅",
                       "少數股東權益", "其他負債合計", "可贖回優先股合計", "不可贖回優先股淨額", "普通股合計", "附加資本", "保留盈餘(累計虧損)",
                       "普通庫存股", "員工持股計劃債務擔保", "未實現收益（虧損）", "其他權益合計"]
    finance1_2_list = ["物業/廠房/設備總額", "累計折舊合計", "長期債務", "資本租賃債務"]
    finance2_1_list = ["銀行應付現金和欠款", "其他盈利資產合計", "淨貸款", "物業/廠房/設備淨總額", "商譽淨額", "無形資產淨額", "長期投資", "其他長期資產合計",
                       "其他資產合計", "應付賬款", "應付/應計", "應計費用", "存款總額", "其他付息負債合計", "短期借貸總額", "長期負債當前應收部分/資本租賃",
                       "其他流動負債合計", "長期債務合計", "遞延所得稅", "少數股東權益", "其他負債合計", "可贖回優先股合計", "不可贖回優先股淨額", "普通股合計",
                       "附加資本", "保留盈餘(累計虧損)", "普通庫存股", "員工持股計劃債務擔保", "未實現收益（虧損）", "其他權益合計"]
    finance2_2_list = ["物業/廠房/設備總額", "累計折舊合計", "長期債務", "資本租賃債務"]
    url = "https://hk.investing.com/instruments/Financials/changereporttypeajax"
    headers = {"User-Agent": user_agent.random,
               "x-requested-with": "XMLHttpRequest"}
    stock_data_annual = {"pair_ID": pair_id, "action": "change_report_type",
                         "report_type": "BAL", "period_type": "Interim"}
    stock_data_interim = {"pair_ID": pair_id, "action": "change_report_type",
                          "report_type": "BAL", "period_type": "Annual"}
    res = requests.post(url, headers=headers, data=stock_data_annual)
    sleep(0.5)
    res2 = requests.post(url, headers=headers, data=stock_data_interim)
    res.encoding = "UTF-8"
    res2.encoding = "UTF-8"
    xml = BeautifulSoup(res.text, "lxml")
    xml2 = BeautifulSoup(res2.text, "lxml")
    try:
        df1 = pd.read_html(str(xml))[0]
    except Exception as e:
        df1 = pd.DataFrame()
        balance = 0
        error = "資產負債表季，"+e
    try:
        df2 = pd.read_html(str(xml2))[0]
    except Exception as e:
        df2 = pd.DataFrame()
        balance = 0
        error = "資產負債表年，"+e
    if not df1.empty and not df2.empty:
        if len(df1.index) == 53:
            bold_list = ["流動資產合計", "總資產", "總流動負債", "總負債",
                         "總權益", "負債及股東權益總計", "已發行普通股合計", "已發行優先股合計"]
            frame1_to_rows = list(dataframe_to_rows(
                df1, index=False, header=False))  # 48rows
            end_date = frame1_to_rows[-1]
            clean_end_date = []
            for date in end_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
                    clean_end_date.append(date_string)
                if str(date) == "nan":
                    clean_end_date.append("")
            clean_end_date[0] = "結束日期："
            frame1_to_rows.pop(1)
            frame1_to_rows.pop(11)
            frame1_to_rows.pop(21)
            frame1_to_rows.pop(28)
            frame1_to_rows.pop(35)
            frame1_to_rows.pop(-1)
            frame1_to_rows.insert(0, clean_end_date)
            frame2_to_rows = list(dataframe_to_rows(
                df2, index=False, header=False))
            end2_date = frame2_to_rows[-1]
            clean_end2_date = []
            for date in end2_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
                    clean_end2_date.append(date_string)
                if str(date) == "nan":
                    clean_end2_date.append("")
            clean_end2_date[0] = "結束日期："
            frame2_to_rows.pop(1)
            frame2_to_rows.pop(11)
            frame2_to_rows.pop(21)
            frame2_to_rows.pop(28)
            frame2_to_rows.pop(35)
            frame2_to_rows.pop(-1)
            frame2_to_rows.insert(0, clean_end2_date)
            for i in frame1_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            for i in frame2_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            if len(balance1_list) != len(frame1_to_rows):
                balance1_list.insert(0, "1")
            final1_rows = []
            final2_rows = []
            for i in balance1_list:
                if i:
                    final1_rows.append(frame1_to_rows[0])
                frame1_to_rows.pop(0)
            for i in balance1_list:
                if i:
                    final2_rows.append(frame2_to_rows[0])
                frame2_to_rows.pop(0)
            for i in final1_rows:
                ws1.append(styled_cells(data=i, key1_list=general1_1_list,
                           key2_list=general1_2_list, key3_list=bold_list, ws=ws1))
            for i in final2_rows:
                ws2.append(styled_cells(data=i, key1_list=general1_1_list,
                           key2_list=general1_2_list, key3_list=bold_list, ws=ws2))
            balance = 0
            error = 0
        elif len(df1.index) == 49:
            bold_list = ["流動資產合計", "總資產", "總流動負債", "總負債",
                         "總權益", "負債及股東權益總計", "已發行普通股合計", "已發行優先股合計"]
            frame1_to_rows = list(dataframe_to_rows(
                df1, index=False, header=False))  # 46rows
            end_date = frame1_to_rows[-1]
            clean_end_date = []
            for date in end_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
                    clean_end_date.append(date_string)
                if str(date) == "nan":
                    clean_end_date.append("")
            clean_end_date[0] = "結束日期："
            frame1_to_rows.pop(2)
            frame1_to_rows.pop(19)
            frame1_to_rows.pop(33)
            frame1_to_rows.pop(-1)
            frame1_to_rows.insert(0, clean_end_date)
            frame2_to_rows = list(dataframe_to_rows(
                df2, index=False, header=False))
            end2_date = frame2_to_rows[-1]
            clean_end2_date = []
            for date in end2_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
                    clean_end2_date.append(date_string)
                if str(date) == "nan":
                    clean_end2_date.append("")
            clean_end2_date[0] = "結束日期："
            frame2_to_rows.pop(2)
            frame2_to_rows.pop(19)
            frame2_to_rows.pop(33)
            frame2_to_rows.pop(-1)
            frame2_to_rows.insert(0, clean_end2_date)
            for i in frame1_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            for i in frame2_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            if len(balance2_list) != len(frame1_to_rows):
                balance2_list.insert(0, "1")
            final1_rows = []
            final2_rows = []
            for i in balance2_list:
                if i:
                    final1_rows.append(frame1_to_rows[0])
                frame1_to_rows.pop(0)
            for i in balance2_list:
                if i:
                    final2_rows.append(frame2_to_rows[0])
                frame2_to_rows.pop(0)
            for i in final1_rows:
                ws1.append(styled_cells(data=i, key1_list=finance1_1_list,
                           key2_list=finance1_2_list, key3_list=bold_list, ws=ws1))
            for i in final2_rows:
                ws2.append(styled_cells(data=i, key1_list=finance1_1_list,
                           key2_list=finance1_2_list, key3_list=bold_list, ws=ws2))
            balance = 0
            error = 0
        elif len(df1.index) == 46:
            bold_list = ["利息收益淨額", "扣除風險準備金後淨利息收入", "稅前淨收益", "稅後淨收益",
                         "計算特殊項目前的淨收益", "淨收入", "扣除特殊項目的普通收入", "稀釋后扣除特殊項目的每股盈利"]
            frame1_to_rows = list(dataframe_to_rows(
                df1, index=False, header=False))  # 43rows
            end_date = frame1_to_rows[-1]
            clean_end_date = []
            for date in end_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
                    clean_end_date.append(date_string)
                if str(date) == "nan":
                    clean_end_date.append("")
            clean_end_date[0] = "結束日期："
            frame1_to_rows.pop(2)
            frame1_to_rows.pop(15)
            frame1_to_rows.pop(30)
            frame1_to_rows.pop(-1)
            frame1_to_rows.insert(0, clean_end_date)
            frame2_to_rows = list(dataframe_to_rows(
                df2, index=False, header=False))
            end2_date = frame2_to_rows[-1]
            clean_end2_date = []
            for date in end2_date:
                if str(date) != "nan":
                    date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
                    clean_end2_date.append(date_string)
                if str(date) == "nan":
                    clean_end2_date.append("")
            clean_end2_date[0] = "結束日期："
            frame2_to_rows.pop(2)
            frame2_to_rows.pop(15)
            frame2_to_rows.pop(30)
            frame2_to_rows.pop(-1)
            frame2_to_rows.insert(0, clean_end2_date)
            for i in frame1_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            for i in frame2_to_rows:
                for index, value in enumerate(i):
                    if i[index] == "-":
                        i[index] = ""
            if len(balance3_list) != len(frame1_to_rows):
                balance3_list.insert(0, "1")
            final1_rows = []
            final2_rows = []
            for i in balance3_list:
                if i:
                    final1_rows.append(frame1_to_rows[0])
                frame1_to_rows.pop(0)
            for i in balance3_list:
                if i:
                    final2_rows.append(frame2_to_rows[0])
                frame2_to_rows.pop(0)
            for i in final1_rows:
                ws1.append(styled_cells(data=i, key1_list=finance2_1_list,
                           key2_list=finance2_2_list, key3_list=bold_list, ws=ws1))
            for i in final2_rows:
                ws2.append(styled_cells(data=i, key1_list=finance2_1_list,
                           key2_list=finance2_2_list, key3_list=bold_list, ws=ws2))
            balance = 0
            error = 0
        else:
            error = "df行數" + str(len(df1.index))
            balance = 1
            # try:
            #     frame1_to_rows = list(dataframe_to_rows(
            #         df1, index=False, header=False))
            #     end_date = frame1_to_rows[-1]
            #     clean_end_date = []
            #     for date in end_date:
            #         if str(date) != "nan":
            #             date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
            #             clean_end_date.append(date_string)
            #         if str(date) == "nan":
            #             clean_end_date.append("")
            #     clean_end_date[0] = "結束日期："
            #     frame1_to_rows.pop(1)
            #     frame1_to_rows.pop(-1)
            #     frame1_to_rows.insert(0, clean_end_date)
            #     frame2_to_rows = list(dataframe_to_rows(
            #         df2, index=False, header=False))
            #     end2_date = frame2_to_rows[-1]
            #     clean_end2_date = []
            #     for date in end2_date:
            #         if str(date) != "nan":
            #             date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
            #             clean_end2_date.append(date_string)
            #         if str(date) == "nan":
            #             clean_end2_date.append("")
            #     clean_end2_date[0] = "結束日期："
            #     frame2_to_rows.pop(1)
            #     frame2_to_rows.pop(-1)
            #     frame2_to_rows.insert(0, clean_end2_date)
            #     balance = 1
            #     error = 0
            # except Exception as e:
            #     frame1_to_rows = []
            #     frame2_to_rows = []
            #     balance = 1
            #     error = e
    return balance, error


"""抓取現金流量表"""
# https://hk.investing.com/instruments/Financials/changereporttypeajax


def get_stock_cashflow(pair_id, choose1_list, ws1, ws2):
    key1_list = ["折舊/遞耗", "攤銷", "遞延稅", "非現金項目", "現金收入", "現金支出", "現金稅金支出", "現金利息支出",
                 "營運資金變動", "資本支出", "其他投資現金流項目合計", "融資現金流項目", "發放現金紅利合計",
                 "股票發行（贖回）淨額", "債務發行（贖回）淨額"]
    cash1_list = choose1_list
    url = "https://hk.investing.com/instruments/Financials/changereporttypeajax"
    headers = {"User-Agent": user_agent.random,
               "x-requested-with": "XMLHttpRequest"}
    stock_data_annual = {"pair_ID": pair_id, "action": "change_report_type",
                         "report_type": "CAS", "period_type": "Interim"}
    stock_data_interim = {"pair_ID": pair_id, "action": "change_report_type",
                          "report_type": "CAS", "period_type": "Annual"}
    res = requests.post(url, headers=headers, data=stock_data_annual)
    sleep(0.5)
    res2 = requests.post(url, headers=headers, data=stock_data_interim)
    res.encoding = "UTF-8"
    res2.encoding = "UTF-8"
    xml = BeautifulSoup(res.text, "lxml")
    xml2 = BeautifulSoup(res2.text, "lxml")
    try:
        df1 = pd.read_html(str(xml))[0]
    except Exception as e:
        df1 = pd.DataFrame()
        cash = 0
        error = "現金流量表季，"+e
    try:
        df2 = pd.read_html(str(xml2))[0]
    except Exception as e:
        df2 = pd.DataFrame()
        cash = 0
        error = "現金流量表年，"+e
    if not df1.empty and not df2.empty:
        bold_list = ["淨收益/起點", "來自經營活動的現金", "來自投資活動的現金", "來自融資活動的現金", "現金變動淨額"]
        frame1_to_rows = list(dataframe_to_rows(
            df1, index=False, header=False))
        end_date = frame1_to_rows[-2]
        end_date_range = frame1_to_rows[-1]
        clean_end_date_range = []
        for i in end_date_range:
            if len(i) <2:
                i = ""
            clean_end_date_range.append(i)
        clean_end_date = []
        for date in end_date:
            if str(date) != "nan":
                date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[3:5]
                clean_end_date.append(date_string)
            if str(date) == "nan":
                clean_end_date.append("")
        clean_end_date[0] = "結束日期："
        frame1_to_rows.pop(2)
        frame1_to_rows.pop(12)
        frame1_to_rows.pop(15)
        frame1_to_rows.pop(-1)
        frame1_to_rows.pop(-1)
        frame1_to_rows.insert(0, clean_end_date_range)
        frame1_to_rows.insert(0, clean_end_date)
        frame2_to_rows = list(dataframe_to_rows(
            df2, index=False, header=False))
        end2_date = frame2_to_rows[-2]
        end2_date_range = frame2_to_rows[-1]
        clean_end2_date_range = []
        for i in end2_date_range:
            if len(i) <2:
                i = ""
            clean_end2_date_range.append(i)
        clean_end2_date = []
        for date in end2_date:
            if str(date) != "nan":
                date_string = str(date)[0:4]+"/"+str(date)[str(date).find("/")+1:]+"/"+str(date)[4:6]
                clean_end2_date.append(date_string)
            if str(date) == "nan":
                clean_end2_date.append("")
        clean_end2_date[0] = "結束日期："
        frame2_to_rows.pop(2)
        frame2_to_rows.pop(12)
        frame2_to_rows.pop(15)
        frame2_to_rows.pop(-1)
        frame2_to_rows.pop(-1)
        frame2_to_rows.insert(0, clean_end2_date_range)
        frame2_to_rows.insert(0, clean_end2_date)
        for i in frame1_to_rows:
            for index, value in enumerate(i):
                if i[index] == "-":
                    i[index] = ""
        for i in frame2_to_rows:
            for index, value in enumerate(i):
                if i[index] == "-":
                    i[index] = ""
        if len(cash1_list) != len(frame1_to_rows):
            cash1_list.insert(0, "1")
            cash1_list.insert(0, "1")
        final1_rows = []
        final2_rows = []
        for i in cash1_list:
            if i:
                final1_rows.append(frame1_to_rows[0])
                final2_rows.append(frame2_to_rows[0])
            frame1_to_rows.pop(0)
            frame2_to_rows.pop(0)
        for i in final1_rows:
            ws1.append(styled_cells(data=i, key1_list=key1_list,
                       key3_list=bold_list, ws=ws1))
        for i in final2_rows:
            ws2.append(styled_cells(data=i, key1_list=key1_list,
                       key3_list=bold_list, ws=ws2))
        cash = 0
        error = 0
    return cash, error


"""抓取比率"""
# https://hk.investing.com/equities/nordea-bank-finland-ratios


def get_stock_ratio(url, ws):
    if "?cid=" in url:
        f_index = url.find("?cid=")
        l_index = url[f_index:]
        url = url[:f_index]+"-ratios"+l_index
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
    else:
        url = url+"-ratios"
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
    pdf = pd.read_html(str(xml))[0]
    pdf1 = pd.read_html(str(xml))[1]
    dfindexs = len(pdf.index)
    df1indexs = len(pdf1.index)
    if dfindexs > 60:
        df = pdf
    elif df1indexs > 60:
        df = pdf1
    frame_to_rows = list(dataframe_to_rows(df, index=False, header=False))
    frame_to_rows.pop(0)
    frame_to_rows.pop(0)
    frame_to_rows.pop(7)
    frame_to_rows[7] = ["獲利率：過去十二個月 vs 5年平均盈利比較", "TTM(%)", "5年平均(%)"]
    frame_to_rows.pop(17)
    frame_to_rows.pop(25)
    frame_to_rows[25] = ["管理績效：過去十二個月 vs 5年平均盈利比較", "TTM(%)", "5年平均(%)"]
    frame_to_rows.pop(33)
    frame_to_rows.pop(41)
    frame_to_rows.pop(46)
    frame_to_rows.pop(52)
    frame_to_rows.insert(45, [""])
    for i in frame_to_rows:
        ws.append(i)


"""抓取股利"""
# https://hk.investing.com/equities/apple-computer-inc


def get_dividend(url, ws):
    if "?cid=" in url:
        f_index = url.find("?cid=")
        l_index = url[f_index:]
        url = url[:f_index]+"-dividends"+l_index
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
    else:
        url = url+"-dividends"
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
    finaldf = pd.DataFrame()
    df = pd.read_html(str(xml))
    for i in df:
        if "除息日" in i.columns:
            # if "除息日" in str(i.columns[0]):
            finaldf = i
    frame_to_rows = list(dataframe_to_rows(finaldf, index=False, header=False))
    for i in frame_to_rows:
        ws.append(i)


"""抓取財報"""
# https://hk.investing.com/equities/apple-computer-inc


def get_financial(url, ws):
    if "?cid=" in url:
        f_index = url.find("?cid=")
        l_index = url[f_index:]
        url = url[:f_index]+"-earnings"+l_index
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
        df = pd.read_html(str(xml))
    else:
        url = url+"-earnings"
        res = requests.get(url, headers=headers)
        res.encoding = "UTF-8"
        xml = BeautifulSoup(res.text, "lxml")
        df = pd.read_html(str(xml))
        finaldf = pd.DataFrame()
    for i in df:
        if "公佈日期" in i.columns:
            finaldf = i
    frame_to_rows = list(dataframe_to_rows(finaldf, index=False, header=False))
    for i in frame_to_rows:
        ieps = str(i[2]) + str(i[3])
        irevenue = str(i[4]) + str(i[5])
        newi = [i[0], i[1], ieps, irevenue]
        ws.append(newi)


"""抓取關鍵字查詢"""
# https://hk.investing.com/search/service/SearchInnerPage


def get_keypoint_stocklist(word):
    search_list = []
    url = "https://hk.investing.com/search/service/SearchInnerPage"
    headers = {"User-Agent": user_agent.random,
               "x-requested-with": "XMLHttpRequest"}
    stock_data = {"search_text": word, "tab": "quotes", "isFilter": True}
    res = requests.post(url, headers=headers, data=stock_data)
    res.encoding = "UTF-8"
    res = res.json()["quotes"]
    for i in res:
        if "股票" in i["type"]:
            name = i["name"].replace(" ", "-")
            location = i["type"].replace(" ", "")
            search_list.append(
                (name+location, "https://hk.investing.com"+i["link"], +i["pairId"]))
    return search_list


"""抓取個股"""


def create_stockdata(i, sdate, edate, f1_1_list, f1_2_list, f1_3_list, f2_1_list, f2_2_list, f2_3_list, f3_1_list, download_path):
    createday = str(date.today())
    errorname = createday+"error_list.xlsx"
    errorfile = Path(download_path) / errorname
    stockpair = ""
    if errorfile.is_file():
        error_wb = load_workbook(errorfile)
        error_ws = error_wb.active
        error_ws.title = "異常個股清單"
    else:
        error_wb = Workbook()
        error_ws = error_wb.active
    wb = Workbook()
    ws1 = wb.worksheets[0]
    ws2 = wb.create_sheet("歷史價格")
    ws3 = wb.create_sheet("財務摘要-季報")
    ws4 = wb.create_sheet("財務摘要-年報")
    ws5 = wb.create_sheet("損益表-季報")
    ws6 = wb.create_sheet("損益表-年報")
    ws7 = wb.create_sheet("資產負債表-季報")
    ws8 = wb.create_sheet("資產負債表-年報")
    ws9 = wb.create_sheet("現金流量表-季報")
    ws10 = wb.create_sheet("現金流量表-年報")
    ws11 = wb.create_sheet("比率")
    ws12 = wb.create_sheet("股息")
    ws13 = wb.create_sheet("財報")
    ws1.title = "綜觀"
    ws1_write_column(ws1)
    ws2_write_column(ws2)
    ws3to4_write_column(ws1=ws3, ws2=ws4)
    ws11_write_column(ws11)
    ws12_write_column(ws12)
    ws13_write_column(ws13)
    pair_id = i[2]
    try:
        profitloss_num, profiloss_error, save_type = get_profitandloss(
            pair_id=pair_id, choose1_list=f1_1_list, choose2_list=f1_2_list, choose3_list=f1_3_list, ws1=ws5, ws2=ws6)  # 損益表
        balance_num = 0
        balance_error = 0
        cash_num = 0
        cash_error = 0
        print("成功")
    except Exception as e:
        profiloss_error = e
        profitloss_num = 0
        balance_num = 0
        balance_error = 0
        cash_num = 0
        cash_error = 0
        save_type = "其它"
        print("失敗")
    print("完成損益表", profiloss_error)
    sleep(0.5)
    if not profiloss_error:
        try:
            balance_num, balance_error = get_stock_balance(
                pair_id=pair_id, choose1_list=f2_1_list, choose2_list=f2_2_list, choose3_list=f2_3_list, ws1=ws7, ws2=ws8)  # 資產負債表
            sleep(0.5)
        except Exception as e:
            balance_error = e
            balance_num = 0
        print("完成資產表", balance_error)
        try:
            cash_num, cash_error = get_stock_cashflow(
                pair_id=pair_id, choose1_list=f3_1_list, ws1=ws9, ws2=ws10)  # 現金流量表
            sleep(0.5)
        except Exception as e:
            cash_error = e
            cash_num = 0
        print("完成現金表", cash_error)
        try:
            get_stock_ratio(url=i[1], ws=ws11)  # 比率
            radio_error = 0
            sleep(1)
        except Exception as e:
            radio_error = e
        print("完成比率", radio_error)
        try:
            stockpair = get_stock_inventory(url=i[1], ws=ws1)  # 綜觀
            inventory_error = 0
            sleep(1)
        except Exception as e:
            inventory_error = e
            stockpair = ""
        print("完成綜觀", inventory_error)
        try:
            get_stock_profile(url=i[1], ws=ws1)  # 簡介
            profile_error = 0
            sleep(1)
        except Exception as e:
            profile_error = e
        print("完成簡介", profile_error)
        try:
            get_stock_history(pair_id=pair_id, start_date=sdate,
                              end_date=edate, ws=ws2)  # 歷史資料
            history_error = 0
            sleep(1)
        except Exception as e:
            history_error = e
        print("完成歷史資料", history_error)
        try:
            get_stock_financials(pair_id=pair_id, ws1=ws3, ws2=ws4)  # 財務摘要
            financials_error = 0
            sleep(0.5)
        except Exception as e:
            financials_error = e
        print("完成財務摘要", financials_error)
        try:
            get_dividend(url=i[1], ws=ws12)  # 股利
            dividend_error = 0
            sleep(0.5)
        except Exception as e:
            dividend_error = e
        print("完成股利", dividend_error)
        try:
            get_financial(url=i[1], ws=ws13)
            financial2_error = 0
            sleep(0.5)
        except Exception as e:
            financial2_error = e
        print("完成財報", financial2_error)
    name = stockpair+i[0].replace("/", "").replace("-", "").replace(" ", "").replace(
        "*", "").replace("<", "").replace(">", "").replace("?", "")+createday+".xlsx"
    format_count = profitloss_num+balance_num+cash_num
    if format_count > 1:
        newname = "全格式"+name
    else:
        newname = name
    if save_type == "一般":
        store_path = download_path / "一般業" / newname
    elif save_type == "保險":
        store_path = download_path / "保險業" / newname
    elif save_type == "銀行":
        store_path = download_path / "銀行業" / newname
    elif save_type == "其它":
        store_path = download_path / "其它業" / newname
    for sheet in wb:
        autowidth(sheet)
    if profiloss_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "損益表 - "+str(profiloss_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif balance_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "資產負債表 - "+str(balance_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif cash_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "現金流量表 - "+str(cash_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif radio_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "比率 - "+str(radio_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif inventory_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "綜觀 - "+str(inventory_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif profile_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "簡介 - "+str(profile_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif history_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "歷史資料 - "+str(history_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif financials_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "財務摘要 - "+str(financials_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif dividend_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "股利 - "+str(dividend_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    elif financial2_error:
        error_write = [i[0], i[1], i[2], str(
            datetime.now()), "財報 - "+str(financial2_error)]
        error_ws.append(error_write)
        error_wb.save(errorfile)
    else:
        wb.save(store_path)
        error_wb.save(errorfile)
    # return profiloss_error#TODO這邊要做錯誤分類回傳

# get_dividend(url="https://hk.investing.com/equities/taiwan-semicon")
# get_profitandloss(pair_id="1012665")
# get_dividend(url="https://hk.investing.com/equities/facebook-inc")
# get_stock_financials(pair_id="26490")
# f1 = ['總收入', '收入', '其他收入合計', '稅收成本合計', '毛利', '經營開支總額', '銷售/一般/管理費用合計', '研發', '折舊/攤銷', '利息開支(收入)-營運淨額', '例外開支(收入)', '其 他運營開支總額', '營業收入', '利息收入（開支）- 非營運淨額', '出售資產收入（虧損）', '其他，淨額', '稅前淨收益', '備付所得稅', '稅後淨收益', '少數股東權益', '附屬公司權益', '美國公認會計準則調整', '計算特殊項目前的淨收益', '特殊項目合計', '淨收入', '淨收入調整總額', '扣除特殊項目的普通收入', '稀釋調整', '稀釋后淨收入', '稀釋后加權平均股', '稀釋后扣除特殊項目的每股盈利', '每股股利 – 普通股首次發行', '稀釋后每股標準盈利']
# f2 = ['總收入', '保費收入合計', '投資收益淨額', '變現收益（虧損）', '其他收入合計', '經營開支總額', '虧損、福利和修訂合計', '购置成本攤銷', '銷售/一般/管理費用合計', '折舊/攤銷', '利息開支（收入）- 營運淨額', '例外開支（收入）', '其他運營開支總額', '營業收入', '利息收入（開支）- 非營運淨額', '出售資產收入（虧損）', '其他，淨額', '稅前淨收益', ' 備付所得稅', '稅後淨收益', '少數股東權益', '附屬公司權益', '美國公認會計準則調整', '計算特殊項目前的淨收益', '特殊項目合計', '淨收入', '淨收入調整總額', '扣除特殊項目的普 通收入', '稀釋調整', '稀釋后淨收入', '稀釋后加權平均股', '稀釋后扣除特殊項目的每股盈利', '每股股利 – 普通股首次發行', '稀釋后每股標準盈利']
# f3 = ['利息收益淨額', '銀行利息收入', '利息開支總額', '風險準備金', '扣除風險準備金後淨利息收入', '銀行非利息收入', '銀行非利息開支', '稅前淨收益', '備付所得稅', '稅後淨收益', '少數股東權益', '附屬公司權益', '美國公認會計準則調整', '計算特殊項目前的淨收益', '特殊項目合計', '淨收入', '淨收入調整總額', '扣除特殊項目的普通收入', '稀釋調整', '稀釋后淨收入', '稀釋后加權平均股', '稀釋后扣除特殊項目的每股盈利', '每股股利 – 普通股首次發行', '稀釋后每股標準盈利']
# c1 = ['流動資產合計', '現金和短期投資', '現金', '現金和現金等價物', '短期投資', '淨應收款合計', '淨交易應收款合計', '庫存合計', '預付費用', '其他流動資產合計', '總資產', '物業/廠房/設備淨總額', '物業/廠房/設備總額', '累計折舊合計', '商譽淨額', '無形資產淨額', '長期投資', '長期應收票據', '其他長期資產合計', '其他資產合計', '總流動負債', '應付賬款', '應付/應計', '應計費用', '應付票據/短期債務', '長期負債當前應收部分/資本租賃', '其他流動負債合計', '總負債', '長期債務合計', '長期債務', '資本租賃債務', '遞延所得稅', '少數股東權益', '其他負債合計', '總權益', '可贖回優先股合計', '不可贖回優先股淨額', '普通股合計', '附加資本', '保留盈餘(累計虧損)', '普通庫存股', '員工持股計劃債務擔保', '未實現收益（虧損）', '其他權益合計', '負債及股東權益總計', '已發行普通股合計', '已發行優先股合計']
# wb = Workbook()
# ws1 = wb.worksheets[0]
# ws2 = wb.create_sheet("歷史價格")
# ws3 = wb.create_sheet("財務摘要-季報")
# ws4 = wb.create_sheet("財務摘要-年報")
# ws5 = wb.create_sheet("損益表-季報")
# ws6 = wb.create_sheet("損益表-年報")
# ws7 = wb.create_sheet("資產負債表-季報")
# ws8 = wb.create_sheet("資產負債表-年報")
# ws9 = wb.create_sheet("現金流量表-季報")
# ws10 = wb.create_sheet("現金流量表-年報")
# ws11 = wb.create_sheet("比率")
# ws12 = wb.create_sheet("股息")
# ws13 = wb.create_sheet("財報")
# ws1.title = "綜觀"
# ws1_write_column(ws1)
# ws2_write_column(ws2)
# ws3to4_write_column(ws1=ws3,ws2=ws4)
# get_keypoint_stocklist(word="微軟")
# get_stock_profile(url="https://hk.investing.com/equities/microsoft-corp?cid=1184309",ws=ws1)
# get_stock_history(pair_id="26490",start_date="2021/04/20",end_date="2021/04/20")
# get_stock_financials(pair_id="1184309",ws1=ws3,ws2=ws4)
# get_financial(url="https://hk.investing.com/equities/facebook-inc")
# try:
#     profitloss_num,profiloss_error = get_profitandloss(pair_id=pair_id,choose1_list=f1,choose2_list=f2,choose3_list=f3,ws1=ws5,ws2=ws6) # 損益表
# except Exception as e:
#     profitloss_num = 0
#     profiloss_error = 1
# wb.save("test.xlsx")
