from tkinter import ttk,filedialog
from tkinter.scrolledtext import ScrolledText
from win32com.client import Dispatch
from pathlib import Path
from time import perf_counter,perf_counter
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from typing import Any
from icon import img
import tkinter as tk
import threading
import pythoncom
import time
import copy
import base64
import babel.numbers
import os

""" 合併EXCEL範例檔案 """

def write_pathconfig(type: str, value: str) -> None:

    """ 寫入存檔設定 """

    tempfile = Path.cwd() / "exceltemp.xlsx"
    wb = load_workbook(tempfile)
    ws = wb.active
    if type == "source":
        ws.cell(row=1, column=2, value=value)
    elif type == "destination":
        ws.cell(row=2, column=2, value=value)
    wb.save(tempfile)

def align_center(root: Any, width: int, height: int) -> None:

    """ GUI視窗設定 """

    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    size = "%dx%d+%d+%d" % (width, height,
                            (screenwidth - width)/2,
                            (screenheight - height)/2)
    root.geometry(size)

def final1and2_excel_event() -> None:

    """ 新個股資料點擊事件 """

    filename = filedialog.askopenfilename(parent=window, initialdir="C:/",
                                          filetypes = (("excel files", "*.xlsx"),
                                                       ("all files", "*.*")))
    EXTEND_FILE_TEXT.config(state="normal")
    EXTEND_FILE_TEXT.delete(1.0,"end")
    EXTEND_FILE_TEXT.insert(1.0, filename)
    EXTEND_FILE_TEXT.config(state="disable")

def key30_excel_event() -> None:

    """ 關鍵30點擊事件 """

    filename = filedialog.askopenfilename(parent=window,initialdir="C:/",
                                          filetypes = (("excel files", "*.xlsx"),
                                                       ("all files", "*.*")))
    KEY30_TEXT.config(state="normal")
    KEY30_TEXT.delete(1.0, "end")
    KEY30_TEXT.insert(1.0, filename)
    KEY30_TEXT.config(state="disable")

def source_event() -> None:

    """ 來源檔案路徑點擊事件 """

    filename = filedialog.askdirectory(parent=window)
    if not filename:
        filename = "C:/"
    SOURCE_PATH_TEXT.config(state="normal")
    SOURCE_PATH_TEXT.delete(1.0, "end")
    SOURCE_PATH_TEXT.insert(1.0, filename)
    SOURCE_PATH_TEXT.config(state="disable")
    write_pathconfig(type="source", value=filename)

def destination_event() -> None:

    """ 目的檔案路徑點擊事件 """

    filename = filedialog.askdirectory(parent=window)
    if not filename:
        filename = "C:/"
    DESTINATION_PATH_TEXT.config(state="normal")
    DESTINATION_PATH_TEXT.delete(1.0, "end")
    DESTINATION_PATH_TEXT.insert(1.0, filename)
    DESTINATION_PATH_TEXT.config(state="disable")
    write_pathconfig(type="destination", value=filename)

def just_open(filename: Any) -> None:

    """ 模擬手動開啟EXCEL """

    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()

def copy_final1and2_xlsx(examplewb: Any, source_wb: Any, destination_wb: Any) -> None:

    """ 新個股資料複製值與格式設定 """

    wb = load_workbook(examplewb)
    wb2 = load_workbook(source_wb)
    wb_sheetnames = list(wb.sheetnames)[13:]
    for sheetname in wb_sheetnames:
        ws = wb[sheetname]
        ws2 = wb2.create_sheet(sheetname)
        for i, row in enumerate(ws.iter_rows()):
            ws2.row_dimensions[i+1].height = ws.row_dimensions[i+1].height
            for j, cell in enumerate(row):
                ws2.column_dimensions[get_column_letter(j+1)].width = ws.column_dimensions[get_column_letter(j+1)].width
                ws2.cell(row=i + 1, column=j + 1, value=cell.value)
                # 設定單元格格式
                source_cell = ws.cell(i+1, j+1)
                target_cell = ws2.cell(i+1, j+1)
                target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)
    wb2.save(destination_wb)
    wb.close()
    wb2.close()

def copy_key30_xlsx(row_: int, col_: int, source_wb: str, destination_ws: Any) -> int:

    """ 關鍵30複製值與格式設定 """

    wb = load_workbook(source_wb, data_only=True)
    try:
        ws = wb["final2"]
    except:
        ws = ""
    if not ws:
        return False
    for i, column in enumerate(ws.iter_cols(min_col=2, max_col=2, min_row=1, max_row=ws.max_row)):
        destination_ws.row_dimensions[i+1].height = ws.row_dimensions[i+1].height
        for j, cell in enumerate(column):
            destination_ws.column_dimensions[get_column_letter(j+1)].width = ws.column_dimensions[get_column_letter(j+1)].width
            destination_ws.cell(row=row_+i+1, column=col_+j+1, value=cell.value)
            # 設定單元格格式
            target_cell =destination_ws.cell(row_+i+1, col_+j+1)
            if cell.has_style:
                target_cell._style = copy.copy(cell._style)
                target_cell.font = copy.copy(cell.font)
                target_cell.border = copy.copy(cell.border)
                target_cell.fill = copy.copy(cell.fill)
                target_cell.number_format = copy.copy(cell.number_format)
                target_cell.protection = copy.copy(cell.protection)
                target_cell.alignment = copy.copy(cell.alignment)
    row_ += 1
    wb.close()
    return row_

def copy_newkey30_xlsx(examplewb: Any, source_wb: Any, destination_wb: Any) -> None:

    """ 新關鍵30複製值與格式設定 """

    wb = load_workbook(examplewb)
    wb2 = load_workbook(source_wb)
    wb_sheetnames = list(wb.sheetnames)[1:]#13
    for sheetname in wb_sheetnames:
        ws = wb[sheetname]
        ws2 = wb2.create_sheet(sheetname)
        for i, row in enumerate(ws.iter_rows()):
            ws2.row_dimensions[i+1].height = ws.row_dimensions[i+1].height
            for j, cell in enumerate(row):
                ws2.column_dimensions[get_column_letter(j+1)].width = ws.column_dimensions[get_column_letter(j+1)].width
                ws2.cell(row=i + 1, column=j + 1, value=cell.value)
                # 設定單元格格式
                source_cell = ws.cell(i+1, j+1)
                target_cell = ws2.cell(i+1, j+1)
                target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)
    wb2.save(destination_wb)
    wb.close()
    wb2.close()

def general_final1and2_event() -> None:

    """ 新個股生成 """

    EXCEL_TIMER_LB.place_forget()
    sample_excel = EXTEND_FILE_TEXT.get("1.0", "end-1c")
    if not sample_excel:
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "請先選擇正確合併檔案."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        return False
    example_excel = load_workbook(sample_excel)
    if len(list(example_excel.sheetnames)) <= 13:
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "該合併檔案無新工作簿."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        example_excel.close()
        return False
    source_pwd = Path(SOURCE_PATH_TEXT.get("1.0", "end-1c"))
    source_pwd_folder = os.listdir(source_pwd)
    if not source_pwd_folder:
        STEXT.config(state="normal")
        STEXT.insert(tk.END,"資料來源路徑沒有個股資料."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        return False
    destination_pwd = Path(DESTINATION_PATH_TEXT.get("1.0", "end-1c"))
    for dir in source_pwd.iterdir():
        full_path = dir
        filename = dir.name
        save_path = destination_pwd / filename
        copy_final1and2_xlsx(examplewb=sample_excel, source_wb=full_path, destination_wb=save_path)
    STEXT.config(state="normal")
    nowtime = time.localtime()
    time_string = time.strftime("%Y-%m-%d %H:%M:%S", nowtime) # 字串
    STEXT.insert(tk.END, time_string+" ----- 合併範例資料生成完畢."+"\n")
    STEXT.see(tk.END)
    STEXT.config(state="disable")

def general_key30_event() -> None:

    """ 關鍵30線呈事件 """

    key30_td = threading.Thread(target=generator_key30_event)
    key30_td.setDaemon(True)
    key30_td.start()

def generator_key30_event() -> None:

    """ 關鍵30生成 """

    pythoncom.CoInitialize()
    source_pwd = Path(SOURCE_PATH_TEXT.get("1.0", "end-1c"))
    source_pwd_folder = os.listdir(source_pwd)
    if not source_pwd_folder:
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "資料來源路徑沒有個股資料."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        return False
    nowtime = time.localtime()
    time_string = time.strftime("%Y-%m-%d %H:%M:%S", nowtime)
    save_timestraing = time.strftime("%Y-%m-%d", nowtime)
    save_path = save_timestraing+"-關鍵30.xlsx"
    destination_pwd = Path(DESTINATION_PATH_TEXT.get("1.0", "end-1c")) / save_path
    if destination_pwd.is_file():
        STEXT.config(state="normal")
        STEXT.insert(tk.END,time_string+" ----- "+"已有生成關鍵30資料."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        return False
    key30wb = Workbook()
    key30ws = key30wb.active
    key30ws.title = "關鍵30指標"
    # 先寫入A欄資料
    first_wbpath = list(source_pwd.iterdir())[0]
    just_open(filename=first_wbpath)
    first_wbname = first_wbpath.name
    firstwb = load_workbook(first_wbpath,data_only=True)
    try:
        firstws = firstwb["final2"]
        first_value_list = [ str(firstws.cell(row=i+1,column=1).value) for i in range(firstws.max_row) ]
        key30ws.append(first_value_list)
        firstwb.close()
    except:
        STEXT.config(state="normal")
        nowtime = time.localtime()
        time_string = time.strftime("%Y-%m-%d %H:%M:%S", nowtime)
        save_timestraing = time.strftime("%Y-%m-%d", nowtime)
        STEXT.insert(tk.END, time_string+" ----- "+first_wbname+"未找到final2工作簿."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
    counter = len(list(source_pwd.iterdir()))*2
    nowrtime = time.time()
    finaltime = nowrtime + counter
    time_struct = time.localtime(finaltime) # 時間元組
    time_string = time.strftime("%Y-%m-%d %H:%M", time_struct) # 字串
    EXCEL_TIMER_LB.configure(text=time_string)
    EXCEL_TIMER_LB.place(relx=0.45,rely=0.475)
    base_row = 1
    base_column = 0
    for dir in source_pwd.iterdir():
        start = perf_counter()
        full_path = dir
        filename = dir.name
        try:
            just_open(filename=full_path)
        except Exception as e:
            print("開檔案失敗", e)
        try:
            base_row=  copy_key30_xlsx(row_=base_row,col_=base_column,source_wb=full_path,destination_ws=key30ws)
        except Exception as e:
            print("複製資料失敗", e)
        STEXT.config(state="normal")
        nowtime = time.localtime()
        time_string = time.strftime("%Y-%m-%d %H:%M:%S", nowtime)
        save_timestraing = time.strftime("%Y-%m-%d", nowtime)
        STEXT.insert(tk.END,time_string+" ----- "+filename+"生成完畢."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        print(f"Cost: {perf_counter() - start}")
        if not base_row:
            nowtime = time.localtime()
            time_string = time.strftime("%Y-%m-%d %H:%M:%S", nowtime)
            save_timestraing = time.strftime("%Y-%m-%d", nowtime)
            STEXT.config(state="normal")
            STEXT.insert(tk.END,time_string+" ----- "+filename+"未找到final2工作簿."+"\n")
            STEXT.see(tk.END)
            STEXT.config(state="disable")
            return False
    key30wb.save(destination_pwd)
    nowtime = time.localtime()
    time_string = time.strftime("%Y-%m-%d %H:%M:%S", nowtime)
    save_timestraing = time.strftime("%Y-%m-%d", nowtime)
    STEXT.config(state="normal")
    STEXT.insert(tk.END,time_string+" ----- 關鍵30資料生成完畢."+"\n")
    STEXT.see(tk.END)
    STEXT.config(state="disable")
    pythoncom.CoUninitialize()

def general_newkey30_event() -> None:

    """ 新關鍵30生成 """

    EXCEL_TIMER_LB.place_forget()
    key30_path = KEY30_TEXT.get("1.0", "end-1c")
    if not key30_path:
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "請先選擇正確合併檔案."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        return False
    example_excel = load_workbook(key30_path)
    if len(list(example_excel.sheetnames)) <= 1:
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "該合併檔案無新工作簿."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        example_excel.close()
        return False
    source_pwd = Path(SOURCE_PATH_TEXT.get("1.0", "end-1c"))
    source_pwd_folder = os.listdir(source_pwd)
    if not source_pwd_folder:
        STEXT.config(state="normal")
        STEXT.insert(tk.END,"資料來源路徑沒有個股資料."+"\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        return False
    destination_pwd = Path(DESTINATION_PATH_TEXT.get("1.0", "end-1c"))
    for dir in source_pwd.iterdir():
        full_path = dir
        filename = "新"+dir.name
        save_path = destination_pwd / filename
        copy_newkey30_xlsx(examplewb=key30_path,source_wb=full_path,destination_wb=save_path)
    STEXT.config(state="normal")
    nowtime = time.localtime()
    time_string = time.strftime("%Y-%m-%d %H:%M:%S", nowtime) # 字串
    STEXT.insert(tk.END,time_string+" ----- 合併關鍵資料生成完畢."+"\n")
    STEXT.see(tk.END)
    STEXT.config(state="disable")

# path configure
path = Path.cwd()

# temp path configure
tempfile = Path.cwd() / "exceltemp.xlsx"
if tempfile.is_file():
    wb = load_workbook(tempfile)
    ws = wb.active
    source_path = ws.cell(row=1, column=2).value
    destination_path = ws.cell(row=2, column=2).value
    wb.close()
else:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="source")
    ws.cell(row=2, column=1, value="destination")
    ws.cell(row=1, column=2, value="C:/")
    ws.cell(row=2, column=2, value="C:/")
    source_path = ws.cell(row=1, column=2).value
    destination_path = ws.cell(row=2, column=2).value
    wb.save(tempfile)

# gui configure
tmp = open("tmp.ico", "wb+")
tmp.write(base64.b64decode(img))
tmp.close()

# ico_path = path / "favicon.ico"
window = tk.Tk()
window.iconbitmap("tmp.ico")
window.title('investing_data_v1')
window.resizable(False, False)
windw_style = ttk.Style(window)
windw_style.theme_use("clam")
windw_style.configure("red.Horizontal.TProgressbar", foreground="green", background="green")
os.remove("tmp.ico")

align_center(window, 1000, 600)
window.update_idletasks()

EXTEND_FILE_LB = tk.Label(window,text="合併範例檔案： ", font=("新細明體", 12), anchor="e")
EXTEND_FILE_LB.place(relx=0.01, rely=0.05, relwidth=0.2)

EXTEND_FILE_BTN = ttk.Button(window,text="選擇檔案",command=final1and2_excel_event)
EXTEND_FILE_BTN.place(relx=0.21, rely=0.045)

EXTEND_FILE_TEXT = tk.Text(window)
EXTEND_FILE_TEXT.config(state="disabled")
EXTEND_FILE_TEXT.place(relx=0.05, rely=0.12, relwidth=0.4, relheight=0.04)

KEY30_LB = tk.Label(window,text="合併關鍵30檔案： ", font=("新細明體", 12), anchor="e")
KEY30_LB.place(relx=0.01, rely=0.2, relwidth=0.2)

KEY30_BTN = ttk.Button(window,text="選擇檔案",command=key30_excel_event)
KEY30_BTN.place(relx=0.21,rely=0.195)

KEY30_TEXT = tk.Text(window)
KEY30_TEXT.config(state="disabled")
KEY30_TEXT.place(relx=0.05,rely=0.27,relwidth=0.4,relheight=0.04)

EXPORT_EXTEND_FILE_BTN = ttk.Button(window,text="生成新個股資料",command=general_final1and2_event)
EXPORT_EXTEND_FILE_BTN.place(relx=0.05,rely=0.4,relwidth=0.25)

EXPORT_KEY30_BTN = ttk.Button(window,text="生成關鍵30資料",command=general_key30_event)
EXPORT_KEY30_BTN.place(relx=0.375,rely=0.4,relwidth=0.25)

EXPORT_NEWKEY30_BTN = ttk.Button(window,text="生成新關鍵30資料",command=general_newkey30_event)
EXPORT_NEWKEY30_BTN.place(relx=0.7,rely=0.4,relwidth=0.25)

SOURCE_PATH_LB = tk.Label(window,text="來源檔案路徑： ",font=("新細明體",12),anchor="e")
SOURCE_PATH_LB.place(relx=0.5,rely=0.05,relwidth=0.2)

SOURCE_PATH_BTN = ttk.Button(window,text="選擇路徑",command=source_event)
SOURCE_PATH_BTN.place(relx=0.71,rely=0.045)

EXCEL_TIMER_LB = tk.Label(window,text="",font=("新細明體",12),anchor="e")
EXCEL_TIMER_LB.place(relx=0.45,rely=0.475)

SOURCE_PATH_TEXT = tk.Text(window)
SOURCE_PATH_TEXT.config(state="normal")
SOURCE_PATH_TEXT.insert(1.0,source_path)
SOURCE_PATH_TEXT.config(state="disabled")
SOURCE_PATH_TEXT.place(relx=0.55,rely=0.12,relwidth=0.4,relheight=0.04)

DESTINATION_PATH_LB = tk.Label(window,text="目的檔案路徑： ",font=("新細明體",12),anchor="e")
DESTINATION_PATH_LB.place(relx=0.5,rely=0.2,relwidth=0.2)

DESTINATION_PATH_BTN = ttk.Button(window,text="選擇路徑",command=destination_event)
DESTINATION_PATH_BTN.place(relx=0.71,rely=0.195)

DESTINATION_PATH_TEXT = tk.Text(window)
DESTINATION_PATH_TEXT.config(state="normal")
DESTINATION_PATH_TEXT.insert(1.0,destination_path)
DESTINATION_PATH_TEXT.config(state="disabled")
DESTINATION_PATH_TEXT.place(relx=0.55,rely=0.27,relwidth=0.4,relheight=0.04)

STEXT = ScrolledText(window,bg="white",selectbackground="blue")
STEXT.config(state="disabled",font=("新細明體",13))
STEXT.place(relx=0.05,rely=0.55,relwidth=0.9,relheight=0.3)


if __name__ == '__main__':
    window.mainloop()