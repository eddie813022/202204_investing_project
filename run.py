
# -*- coding:utf-8 -*-
from configparser import NoOptionError
from tkinter import ttk, filedialog
from tkinter.scrolledtext import ScrolledText
from tkcalendar import DateEntry
from pathlib import Path
from openpyxl import load_workbook
from typing import Callable, Any
from icon import img
import babel.numbers # pyinstaller生成tkinter時的依賴套件
import tkinter as tk
import threading
import base64
import time
import os
from crawler_request import get_pathconfig
from crawler_request import get_exchang_dict
from crawler_request import create_stockdata
from crawler_request import get_stock_list
from crawler_request import get_country_dict
from crawler_request import read_stock_list
from crawler_request import get_dateconfig
from crawler_request import write_dateconfig
from crawler_request import get_keypoint_stocklist
from crawler_request import get_checkboxconfig
from crawler_request import write_pathconfig

""" 圖形化主程式 """

""" 路徑設定 """

path = Path.cwd()

""" GUI初始化設定 """

tmp = open("tmp.ico","wb+")
tmp.write(base64.b64decode(img))
tmp.close()
window = tk.Tk()
# gui favicon
window.iconbitmap("tmp.ico")
window.title('investing_v2.3')
window.resizable(False,False)
# gui theme
windw_style = ttk.Style(window)
windw_style.theme_use('clam')
# progressbar
windw_style.configure("red.Horizontal.TProgressbar", foreground='green', background='green')
os.remove("tmp.ico")

stop_threads = False
country_dict = get_country_dict()
country_list: list[dict[str, str]] = [ i for i in country_dict.keys() ]
country_list.insert(0,"")
search_list: list[str] = []
market_list: list[dict] = []
CATEGORY_LIST = [ {'所有板塊': '-1'}, {'消费类（非周期性）': '8'}, {'交通运输': '10'}, {'公用事业': '22'},
                 {'医疗保健': '18'}, {'基础材料': '7'}, {'工業': '15'}, {'房地產': '23'}, {'服务': '2'},
                 {'材料': '14'}, {'消費者日常用品': '17'}, {'消费类（周期性）': '3'}, {'生产资料': '5'},
                 {'科技': '4'}, {'综合性企业': '12'}, {'能源': '13'}, {'資訊科技': '20'}, {'通訊服務': '21'},
                 {'金融': '19'}, {'非必需消費品': '16'} ]

FRAME1_1_INSIDE_LIST = [ "總收入","收入","其他收入合計","稅收成本合計","毛利","經營開支總額","銷售/一般/管理費用合計",
                        "研發","折舊/攤銷","利息開支（收入）- 營運淨額","例外開支（收入）","其他運營開支總額","營業收入",
                        "利息收入（開支）- 非營運淨額","出售資產收入（虧損）","其他，淨額","稅前淨收益","備付所得稅",
                        "稅後淨收益","少數股東權益","附屬公司權益","美國公認會計準則調整","計算特殊項目前的淨收益","特殊項目合計",
                        "淨收入","淨收入調整總額","扣除特殊項目的普通收入","稀釋調整","稀釋后淨收入","稀釋后加權平均股",
                        "稀釋后扣除特殊項目的每股盈利","每股股利 – 普通股首次發行","稀釋后每股標準盈利" ]

FRAME1_2_INSIDE_LIST = [ "總收入","保費收入合計","投資收益淨額","變現收益（虧損）","其他收入合計","經營開支總額","虧損、福利和修訂合計",
                        "购置成本攤銷","銷售/一般/管理費用合計","折舊/攤銷","利息開支（收入）- 營運淨額","例外開支（收入）","其他運營開支總額",
                        "營業收入","利息收入（開支）- 非營運淨額","出售資產收入（虧損）","其他，淨額","稅前淨收益","備付所得稅",
                        "稅後淨收益","少數股東權益","附屬公司權益","美國公認會計準則調整","計算特殊項目前的淨收益","特殊項目合計",
                        "淨收入","淨收入調整總額","扣除特殊項目的普通收入","稀釋調整","稀釋后淨收入","稀釋后加權平均股",
                        "稀釋后扣除特殊項目的每股盈利","每股股利 – 普通股首次發行","稀釋后每股標準盈利" ]

FRAME1_3_INSIDE_LIST = [ "利息收益淨額","銀行利息收入","利息開支總額","風險準備金","扣除風險準備金後淨利息收入","銀行非利息收入",
                        "銀行非利息開支","稅前淨收益","備付所得稅","稅後淨收益","少數股東權益","附屬公司權益","美國公認會計準則調整",
                        "計算特殊項目前的淨收益","特殊項目合計","淨收入","淨收入調整總額","扣除特殊項目的普通收入","稀釋調整",
                        "稀釋后淨收入","稀釋后加權平均股","稀釋后扣除特殊項目的每股盈利","每股股利 – 普通股首次發行","稀釋后每股標準盈利" ]

FRAME2_1_INSIDE_LIST = [ "流動資產合計","現金和短期投資","現金","現金和現金等價物","短期投資","淨應收款合計","淨交易應收款合計",
                        "庫存合計","預付費用","其他流動資產合計","總資產","物業/廠房/設備淨總額","物業/廠房/設備總額",
                        "累計折舊合計","商譽淨額","無形資產淨額","長期投資","長期應收票據","其他長期資產合計","其他資產合計",
                        "總流動負債","應付賬款","應付/應計","應計費用","應付票據/短期債務","長期負債當前應收部分/資本租賃",
                        "其他流動負債合計","總負債","長期債務合計","長期債務","資本租賃債務","遞延所得稅","少數股東權益",
                        "其他負債合計","總權益","可贖回優先股合計","不可贖回優先股淨額","普通股合計","附加資本","保留盈餘(累計虧損)",
                        "普通庫存股","員工持股計劃債務擔保","未實現收益（虧損）","其他權益合計","負債及股東權益總計",
                        "已發行普通股合計","已發行優先股合計" ]

FRAME2_2_INSIDE_LIST = [ "流動資產合計","總資產","現金","現金和現金等價物","淨應收款合計","預付費用","物業/廠房/設備淨總額","物業/廠房/設備總額","累計折舊合計",
                        "商譽淨額","無形資產淨額","長期投資","應收保險","長期應收票據","其他長期資產合計","遞延保單獲得成本","其他資產合計","總流動負債",
                        "總負債","應付賬款","應付/應計","應計費用","保單負債","應付票據/短期債務","長期負債當前應收部分/資本租賃","其他流動負債合計",
                        "長期債務合計","長期債務","資本租賃債務","遞延所得稅","少數股東權益","其他負債合計","總權益","可贖回優先股合計","不可贖回優先股淨額",
                        "普通股合計","附加資本","保留盈餘(累計虧損)","普通庫存股","員工持股計劃債務擔保","未實現收益（虧損）","其他權益合計",
                        "負債及股東權益總計","已發行普通股合計","已發行優先股合計" ]

FRAME2_3_INSIDE_LIST = [ "流動資產合計","總資產","銀行應付現金和欠款","其他盈利資產合計","淨貸款","物業/廠房/設備淨總額","物業/廠房/設備總額","累計折舊合計",
                        "商譽淨額","無形資產淨額","長期投資","其他長期資產合計","其他資產合計","總流動負債","總負債","應付賬款","應付/應計","應計費用",
                        "存款總額","其他付息負債合計","短期借貸總額","長期負債當前應收部分/資本租賃","其他流動負債合計","長期債務合計","長期債務","資本租賃債務",
                        "遞延所得稅","少數股東權益","其他負債合計","總權益","可贖回優先股合計","不可贖回優先股淨額","普通股合計","附加資本","保留盈餘(累計虧損)",
                        "普通庫存股","員工持股計劃債務擔保","未實現收益（虧損）","其他權益合計","負債及股東權益總計","已發行普通股合計","已發行優先股合計" ]

FRAME3_1_INSIDE_LIST = [ "淨收益/起點","來自經營活動的現金","折舊/遞耗","攤銷","遞延稅","非現金項目","現金收入","現金支出","現金稅金支出",
                        "現金利息支出","營運資金變動","來自投資活動的現金","資本支出","其他投資現金流項目合計","來自融資活動的現金",
                        "融資現金流項目","發放現金紅利合計","股票發行（贖回）淨額","債務發行（贖回）淨額","外匯影響","現金變動淨額",
                        "期初現金結餘","期末現金結餘","自由現金流","自由現金流增長","自由現金流收益率" ]

F2_1_BLUE_BTN_LIST = [ "現金和短期投資","淨應收款合計","庫存合計","預付費用","其他流動資產合計",
                      "物業/廠房/設備淨總額","商譽淨額","無形資產淨額","長期投資",
                      "長期應收票據","其他長期資產合計","其他資產合計","應付賬款",
                      "應付/應計","應計費用","應付票據/短期債務","長期負債當前應收部分/資本租賃",
                      "其他流動負債合計","長期債務合計","遞延所得稅","少數股東權益","其他負債合計",
                      "可贖回優先股合計","不可贖回優先股淨額","普通股合計","附加資本","保留盈餘(累計虧損)",
                      "普通庫存股","員工持股計劃債務擔保","未實現收益（虧損）","其他權益合計" ]

F2_1_GREEN_BTN_LIST = [ "現金","現金和現金等價物","短期投資","淨交易應收款合計","物業/廠房/設備總額","累計折舊合計","長期債務","資本租賃債務" ]

F2_2_BLUE_BTN_LIST = [ "現金","現金和現金等價物","淨應收款合計","預付費用","物業/廠房/設備淨總額","商譽淨額","無形資產淨額","長期投資",
                      "應收保險","長期應收票據","其他長期資產合計","遞延保單獲得成本","其他資產合計","應付賬款","應付/應計","應計費用",
                      "保單負債","應付票據/短期債務","長期負債當前應收部分/資本租賃","其他流動負債合計","長期債務合計","遞延所得稅",
                      "少數股東權益","其他負債合計","可贖回優先股合計","不可贖回優先股淨額","普通股合計","附加資本","保留盈餘(累計虧損)",
                      "普通庫存股","員工持股計劃債務擔保","未實現收益（虧損）","其他權益合計" ]

F2_2_GREEN_BTN_LIST = [ "物業/廠房/設備總額","累計折舊合計","長期債務","資本租賃債務" ]

F2_3_BLUE_BTN_LIST = [ "銀行應付現金和欠款","其他盈利資產合計","淨貸款","物業/廠房/設備淨總額","商譽淨額","無形資產淨額","長期投資","其他長期資產合計",
                      "其他資產合計","應付賬款","應付/應計","應計費用","存款總額","其他付息負債合計","短期借貸總額","長期負債當前應收部分/資本租賃",
                      "其他流動負債合計","長期債務合計","遞延所得稅","少數股東權益","其他負債合計","可贖回優先股合計","不可贖回優先股淨額","普通股合計",
                      "附加資本","保留盈餘(累計虧損)","普通庫存股","員工持股計劃債務擔保","未實現收益（虧損）","其他權益合計" ]

F2_3_GREEN_BTN_LIST = [ "物業/廠房/設備總額","累計折舊合計","長期債務","資本租賃債務" ]

F1_1_BLUE_BTN_LIST = [ "收入","其他收入合計","銷售/一般/管理費用合計","研發","折舊/攤銷","利息開支（收入）- 營運淨額","例外開支（收入）","其他運營開支總額" ]

F1_2_BLUE_BTN_LIST = [ "保費收入合計","投資收益淨額","變現收益（虧損）","其他收入合計","虧損、福利和修訂合計","购置成本攤銷",
                        "銷售/一般/管理費用合計","折舊/攤銷","利息開支（收入）- 營運淨額","例外開支（收入）","其他運營開支總額"]

F1_3_BLUE_BTN_LIST = [ "銀行利息收入","利息開支總額" ]

F3_1_BLUE_BTN_LIST = [ "折舊/遞耗","攤銷","遞延稅","非現金項目","現金收入","現金支出","現金稅金支出","現金利息支出",
                      "營運資金變動","資本支出","其他投資現金流項目合計","融資現金流項目","發放現金紅利合計",
                      "股票發行（贖回）淨額","債務發行（贖回）淨額"]

""" 建立複選框變數清單 """

F1_1_VAR_NAME_lIST = [ "f1_in_f1_va"+str(x+1) for x in range(0,33) ]
F1_2_VAR_NAME_lIST = [ "f1_in_f2_va"+str(x+1) for x in range(0,34) ]
F1_3_VAR_NAME_lIST = [ "f1_in_f3_va"+str(x+1) for x in range(0,24) ]
F2_1_VAR_NAME_lIST = [ "f2_in_f1_va"+str(x+1) for x in range(0,47) ]
F2_2_VAR_NAME_lIST = [ "f2_in_f2_va"+str(x+1) for x in range(0,45) ]
F2_3_VAR_NAME_lIST = [ "f2_in_f3_va"+str(x+1) for x in range(0,42) ]
F3_1_VAR_NAME_lIST = [ "f3_in_f1_va"+str(x+1) for x in range(0,26) ]

""" 建立複選框按鈕變數清單 """
F1_1_CHECK_BTN_LIST = [ "f1_1_check_btn" +str(x+1) for x in range(0,33) ]
F1_2_CHECK_BTN_LIST = [ "f1_2_check_btn" +str(x+1) for x in range(0,34) ]
F1_3_CHECK_BTN_LIST = [ "f1_3_check_btn" +str(x+1) for x in range(0,24) ]
F2_1_CHECK_BTN_LIST = [ "f2_1_check_btn" +str(x+1) for x in range(0,47) ]
F2_2_CHECK_BTN_LIST = [ "f2_2_check_btn" +str(x+1) for x in range(0,45) ]
F2_3_CHECK_BTN_LIST = [ "f2_3_check_btn" +str(x+1) for x in range(0,42) ]
F3_1_CHECK_BTN_LIST = [ "f3_1_check_btn" +str(x+1) for x in range(0,26) ]


def country_select_event() -> None:
    """ 國家選單點擊事件  """
    global market_list
    country_name = COUNTRY_BTN.get()
    if country_name:
        if country_name in country_list:
            country_code = country_dict[country_name]
            STEXT.config(state="normal")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.insert(tk.END, "正在查詢"+country_name+"交易所數量..\n")
            exchange_dict = get_exchang_dict({country_name:country_code})
            MARKET_BTN.set("") # 將選項清空
            market_list = exchange_dict
            market_name_list = [ i for i in market_list ]
            MARKET_BTN["value"] = market_name_list
            MARKET_BTN.current(0)
            category_name_list = [ j for i in CATEGORY_LIST for j, k in i.items() ]
            CATEGORY_BTN["value"] = category_name_list
            CATEGORY_BTN.current(0)
            STEXT.insert(tk.END, country_name+"共有"+str(len(market_list)-1)+"筆交易所.\n")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.see(tk.END)
            STEXT.config(state="disable")
        else:
            STEXT.config(state="normal")
            STEXT.insert(tk.END, "未正確選擇國家\n")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.see(tk.END)
            STEXT.config(state="disable")

def country_enter_envet() -> None:
    """ 個股輸入欄ENTER事件 """
    global search_list
    if search_list:
        search_list.clear()
    word = COUNTRY_BTN.get()
    search_list = get_keypoint_stocklist(word)
    if search_list:
        for i in search_list:
            STEXT.config(state="normal")
            STEXT.insert(tk.END, i[0]+"\n")
            if i != search_list[-1]:
                STEXT.insert(tk.END, "-------------------------------------\n")
            STEXT.see(tk.END)
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.config(state="disable")

def country_btn_event(event: Callable) -> None:
    """ 國家選單線呈事件觸發 """
    country_btn_td = threading.Thread(target=country_select_event)
    country_btn_td.setDaemon(True)
    country_btn_td.start()

def country_btn_enter_event(event: Callable) -> None:
    """ 個股輸入欄ENTER線呈事件觸發 """
    country_select_td = threading.Thread(target=country_enter_envet)
    country_select_td.setDaemon(True)
    country_select_td.start()

def single_loading_event(num: int) -> None:
    """ 個股讀取條事件 """
    PROGRESS["value"] = 0
    if num == 16:
        timer = 6.5
    elif num == 14:
        timer = 7
    for i in range(num):
        if PROGRESS["value"] < PROGRESS["maximum"]:
            PROGRESS["value"] += timer
            window.update()
            time.sleep(1)
    PROGRESS["value"] = 100
    window.update()

def bulk_loading_event(num: int) -> None:
    """ 批量個股讀取條事件 """
    if num < 100:
        addnum = int(100 / num)
        while PROGRESS["value"] < PROGRESS["maximum"]:
            PROGRESS["value"] += addnum
            window.update()
            time.sleep(16)
            if stop_threads:
                PROGRESS["value"] = 100
                break
    elif num > 100:
        addnum = int(num / 100)
        while PROGRESS["value"] < PROGRESS["maximum"]:
            PROGRESS["value"] += 1
            window.update()
            time.sleep(addnum*16)
            if stop_threads:
                PROGRESS["value"] = 100
                break
    PROGRESS["value"] = 100
    window.update()

def start_date_btn_event(event: Callable) -> None:
    """ 開始日期點擊事件 """
    date_ = START_DATE_BTN.get()
    write_dateconfig(date_)

def singlecrawler_btn_event() -> None:
    """ 個股下載線呈事件 """
    single_btn_td = threading.Thread(target=singledownload_event)
    single_btn_td.setDaemon(True)
    single_btn_td.start()

def bulkcrawler_btn_event() -> None:
    """ 批量下載線呈事件 """
    bulk_btn_td = threading.Thread(target=bulkdownload_event)
    bulk_btn_td.setDaemon(True)
    bulk_btn_td.start()

def crawlerstocklist_btn_event() -> None:
    """ 個股清單下載線呈事件 """
    stocklist_btn_td = threading.Thread(target=crawlerstocklist_event)
    stocklist_btn_td.setDaemon(True)
    stocklist_btn_td.start()

def stopcrawler_btn_event() -> None:
    """ 暫停下載線呈事件 """
    stop_btn_td =threading.Thread(target=stopcrawler_event)
    stop_btn_td.setDaemon(True)
    stop_btn_td.start()

def crawlerstocklist_event() -> None:
    """ 個股清單下載點擊事件 """
    PROGRESS["value"] = 0
    STOCK_TIMER_LB.place_forget()
    CRAWLER_STOCK_LIST_BTN.config(state="disable")
    BULK_CRAWLER_BTN.config(state="disable")
    BULK_START_BTN.config(state="disable")
    SINGLE_CRAWLER_BTN.config(state="disable")
    download_path = Path(DOWNLOAD_PATH_TEXT.get("1.0","end-1c"))
    country_name = COUNTRY_BTN.get()
    exchange_name = MARKET_BTN.get()
    category_name = CATEGORY_BTN.get()
    prefilename1 =  country_name+"-"+exchange_name+"-"+category_name+".xlsx"
    prefilename2 = country_name+"-"+exchange_name+".xlsx"
    prefilename3 = country_name+".xlsx"
    if exchange_name and category_name:
        save_name = download_path / prefilename1
    elif exchange_name:
        save_name = download_path / prefilename2
    else:
        save_name = download_path / prefilename3
    try:
        country_code = country_dict[country_name]
    except:
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.insert(tk.END, "請先選取國家及交易所.\n")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        country_code = ""
    if market_list:
        try:
            exchange_code = market_list.get(exchange_name)
            if exchange_code == "-1":
                exchange_code = "a"
        except:
            exchange_code = "a"
    else:
        exchange_code = "a"
    try:
        for i in CATEGORY_LIST:
            if category_name in i.keys():
                category_code = i.get(category_name)
                if category_code == "-1":
                    category_code = "a"
    except:
        category_code = "a"
    if country_code and exchange_code and category_code:
        stock_data = {"cname":country_name,
                    "ccode":country_code,
                    "ename":exchange_name,
                    "ecode":exchange_code,
                    "gname":category_name,
                    "gcode":category_code,
                    "PROGRESS":PROGRESS,
                    "save":save_name}
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.insert(tk.END, f"正在進行{country_name}-{exchange_name}-{category_name}個股清單抓取.\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        stock_totalcount = str(get_stock_list(**stock_data))
        if int(stock_totalcount) <1:
            STEXT.config(state="normal")
            STEXT.insert(tk.END, f"{country_name}-{exchange_name}-{category_name}查無個股資料.\n")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.see(tk.END)
            STEXT.config(state="disable")
        else:
            STEXT.config(state="normal")
            STEXT.insert(tk.END, f"完成{country_name}-{exchange_name}-{category_name}個股清單抓取({stock_totalcount}).\n")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.see(tk.END)
            STEXT.config(state="disable")
    else:
        PROGRESS["value"] = 100
    CRAWLER_STOCK_LIST_BTN.config(state="normal")
    BULK_CRAWLER_BTN.config(state="normal")
    BULK_START_BTN.config(state="normal")
    SINGLE_CRAWLER_BTN.config(state="normal")

def singledownload_event() -> None:
    STOCK_TIMER_LB.place_forget()
    PROGRESS["value"] = 0
    entertext = ""
    CRAWLER_STOCK_LIST_BTN.config(state="disable")
    BULK_CRAWLER_BTN.config(state="disable")
    BULK_START_BTN.config(state="disable")
    SINGLE_CRAWLER_BTN.config(state="disable")
    start = time.perf_counter()
    tempfile = Path.cwd() / "temp.xlsx"
    checkwb = load_workbook(tempfile)
    checkws = checkwb.active
    try:
        f1general_choose_list = [ globals()["f1_in_f1_va"+str(i)].get() for i in range(len(F1_1_VAR_NAME_lIST)) ]
        f1finance_choose_list = [ globals()["f1_in_f2_va"+str(i)].get() for i in range(len(F1_2_VAR_NAME_lIST)) ]
        f1finance2_choose_list = [ globals()["f1_in_f3_va"+str(i)].get() for i in range(len(F1_3_VAR_NAME_lIST)) ]
        f2general_choose_list = [ globals()["f2_in_f1_va"+str(i)].get() for i in range(len(F2_1_VAR_NAME_lIST)) ]
        f2finance_choose_list = [ globals()["f2_in_f2_va"+str(i)].get() for i in range(len(F2_2_VAR_NAME_lIST)) ]
        f2finance2_choose_list = [ globals()["f2_in_f3_va"+str(i)].get() for i in range(len(F2_3_VAR_NAME_lIST)) ]
        f3general_choose_list = [ globals()["f3_in_f1_va"+str(i)].get() for i in range(len(F3_1_VAR_NAME_lIST)) ]
        download_path = Path(DOWNLOAD_PATH_TEXT.get("1.0","end-1c"))
        start_date = START_DATE_BTN.get()
        end_date = END_DATE_BTN.get()
        for i in range(2,35):
            checkws.cell(row=i, column=2, value=f1general_choose_list[i-2])
        for i in range(35,69):
            checkws.cell(row=i, column=2, value=f1finance_choose_list[i-35])
        for i in range(69,93):
            checkws.cell(row=i, column=2, value=f1finance2_choose_list[i-69])
        for i in range(93,140):
            checkws.cell(row=i, column=2, value=f2general_choose_list[i-93])
        for i in range(140,185):
            checkws.cell(row=i, column=2, value=f2finance_choose_list[i-140])
        for i in range(185,227):
            checkws.cell(row=i, column=2, value=f2finance2_choose_list[i-185])
        for i in range(227,253):
            checkws.cell(row=i, column=2, value=f3general_choose_list[i-227])
        checkwb.save(tempfile)
        entertext = STEXT.selection_get()
    except:
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.insert(tk.END, "請輸入個股關鍵字或代號查詢並選取\n")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
    if entertext:
        if entertext in country_list:
            STEXT.config(state="normal")
            STEXT.insert(tk.END, "請先輸入個股關鍵字或代號查詢並選取\n")
            STEXT.see(tk.END)
            STEXT.config(state="disable")
        elif entertext in [ i[0] for i in search_list ]:
            progres_td = threading.Thread(target=single_loading_event, args=(16, ))
            progres_td.setDaemon(True)
            progres_td.start()
            folder1 = download_path / "一般業"
            folder2 = download_path / "保險業"
            folder3 = download_path / "銀行業"
            folder1.mkdir(exist_ok=True)
            folder2.mkdir(exist_ok=True)
            folder3.mkdir(exist_ok=True)
            for i in search_list:
                if entertext == i[0]:
                    print(i[0])
                    STEXT.config(state="normal")
                    STEXT.insert(tk.END, "=====================================\n")
                    STEXT.insert(tk.END, f"開始抓取({entertext})個股.\n")
                    STEXT.insert(tk.END, "=====================================\n")
                    STEXT.see(tk.END)
                    STEXT.config(state="disable")
                    create_stockdata(i=i, sdate=start_date, edate=end_date, f1_1_list=f1general_choose_list,
                                     f1_2_list=f1finance_choose_list, f1_3_list=f1finance2_choose_list,
                                     f2_1_list=f2general_choose_list, f2_2_list=f2finance_choose_list,
                                     f2_3_list=f2finance2_choose_list, f3_1_list=f3general_choose_list,
                                     download_path=download_path)
                    STEXT.config(state="normal")
                    STEXT.insert(tk.END, "=====================================\n")
                    STEXT.insert(tk.END, f"完成抓取{entertext}個股.\n")
                    STEXT.insert(tk.END, "=====================================\n")
                    STEXT.see(tk.END)
                    STEXT.config(state="disable")
    else:
        PROGRESS["value"] = 100
    print(f"Cost: {time.perf_counter() - start}")
    CRAWLER_STOCK_LIST_BTN.config(state="normal")
    BULK_CRAWLER_BTN.config(state="normal")
    BULK_START_BTN.config(state="normal")
    SINGLE_CRAWLER_BTN.config(state="normal")

def bulkdownload_event() -> None:
    """ 批量下載點擊事件 """
    global stop_threads
    stop_threads = False
    PROGRESS["value"] = 0
    tempfile = Path.cwd() / "temp.xlsx"
    checkwb = load_workbook(tempfile)
    checkws = checkwb.active
    try:
        stock_path = Path(STOCK_LIST_PATH_TEXT.get("1.0", "end-1c"))
    except:
        PROGRESS["value"] = 100
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.insert(tk.END, "請先指定個股清單.\n")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
    try:
        wb = load_workbook(stock_path)
        ws = wb.active
        f1general_choose_list = [ globals()["f1_in_f1_va"+str(i)].get() for i in range(len(F1_1_VAR_NAME_lIST)) ]
        f1finance_choose_list = [ globals()["f1_in_f2_va"+str(i)].get() for i in range(len(F1_2_VAR_NAME_lIST)) ]
        f1finance2_choose_list = [ globals()["f1_in_f3_va"+str(i)].get() for i in range(len(F1_3_VAR_NAME_lIST)) ]
        f2general_choose_list = [ globals()["f2_in_f1_va"+str(i)].get() for i in range(len(F2_1_VAR_NAME_lIST)) ]
        f2finance_choose_list = [ globals()["f2_in_f2_va"+str(i)].get() for i in range(len(F2_2_VAR_NAME_lIST)) ]
        f2finance2_choose_list = [ globals()["f2_in_f3_va"+str(i)].get() for i in range(len(F2_3_VAR_NAME_lIST)) ]
        f3general_choose_list = [ globals()["f3_in_f1_va"+str(i)].get() for i in range(len(F3_1_VAR_NAME_lIST)) ]
        download_path = Path(BULK_DOWNLOAD_PATH_TEXT.get("1.0", "end-1c"))
        start_date = START_DATE_BTN.get()
        end_date = END_DATE_BTN.get()
        for i in range(2,35):
            checkws.cell(row=i, column=2, value=f1general_choose_list[i-2])
        for i in range(35,69):
            checkws.cell(row=i, column=2, value=f1finance_choose_list[i-35])
        for i in range(69,93):
            checkws.cell(row=i, column=2, value=f1finance2_choose_list[i-69])
        for i in range(93,140):
            checkws.cell(row=i, column=2, value=f2general_choose_list[i-93])
        for i in range(140,185):
            checkws.cell(row=i, column=2, value=f2finance_choose_list[i-140])
        for i in range(185,227):
            checkws.cell(row=i, column=2, value=f2finance2_choose_list[i-185])
        for i in range(227,253):
            checkws.cell(row=i, column=2, value=f3general_choose_list[i-227])
        checkwb.save(tempfile)
        last_index = int(read_stock_list(stock_path))
        if last_index:
            folder1 = download_path / "一般業"
            folder2 = download_path / "保險業"
            folder3 = download_path / "銀行業"
            folder1.mkdir(exist_ok=True)
            folder2.mkdir(exist_ok=True)
            folder3.mkdir(exist_ok=True)
            counter_num = (ws.max_row - last_index)+1
            counter_time = counter_num * 16
            nowtime = time.time() # 時間戳
            finaltime = nowtime + counter_time
            time_struct = time.localtime(finaltime) # 時間元組
            time_string = time.strftime("%Y-%m-%d %H:%M", time_struct) # 字串
            STOCK_TIMER_LB.configure(text=time_string)
            STOCK_TIMER_LB.place(relx=0.53, rely=0.88)
            STEXT.config(state="normal")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.insert(tk.END, f"預計完成時間為{time_string}.\n")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.see(tk.END)
            STEXT.config(state="disable")
            progres_bk = threading.Thread(target=bulk_loading_event, args=(counter_num,))
            progres_bk.setDaemon(True)
            progres_bk.start()
            for i in range(last_index, ws.max_row+1):
                name = ws.cell(row=last_index, column=1).value
                link = ws.cell(row=last_index, column=2).value
                id = ws.cell(row=last_index, column=3).value
                irow = (name, link, id)
                STEXT.config(state="normal")
                STEXT.insert(tk.END, "=====================================\n")
                STEXT.insert(tk.END, f"開始抓取第{last_index}個，({name})個股.\n")
                STEXT.see(tk.END)
                STEXT.config(state="disable")
                create_stockdata(i=irow, sdate=start_date, edate=end_date, f1_1_list=f1general_choose_list,
                                 f1_2_list=f1finance_choose_list, f1_3_list=f1finance2_choose_list,
                                 f2_1_list=f2general_choose_list, f2_2_list=f2finance_choose_list,
                                 f2_3_list=f2finance2_choose_list, f3_1_list=f3general_choose_list,
                                 download_path=download_path)
                if stop_threads:
                    ws.cell(row=last_index, column=4, value="finished")
                    STEXT.config(state="normal")
                    STEXT.insert(tk.END, "=====================================\n")
                    STEXT.insert(tk.END, f"暫停至第{last_index}個，({name})個股抓取.\n")
                    STEXT.insert(tk.END, "=====================================\n")
                    STEXT.see(tk.END)
                    STEXT.config(state="disable")
                    last_index += 1
                    wb.save(stock_path)
                    STOCK_TIMER_LB.place_forget()
                    break
                ws.cell(row=last_index, column=4, value="finished")
                STEXT.config(state="normal")
                STEXT.insert(tk.END, f"完成抓取第{last_index}個，({name})個股.\n")
                STEXT.insert(tk.END, "=====================================\n")
                STEXT.see(tk.END)
                STEXT.config(state="disable")
                last_index += 1
            wb.save(stock_path)
            PROGRESS["value"] = 100
        else:
            STEXT.config(state="normal")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.insert(tk.END, "該個股清單已抓取完成.\n")
            STEXT.insert(tk.END, "=====================================\n")
            STEXT.see(tk.END)
            STEXT.config(state="disable")
            wb.save(stock_path)
            PROGRESS["value"] = 100
    except Exception as e:
        print(e)
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.insert(tk.END, "請指定正確個股清單.\n")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
        wb.save(stock_path)
    PROGRESS["value"] = 100

def stopcrawler_event() -> None:
    """ 暫停下載點擊事件 """
    global stop_threads
    PROGRESS["value"] = 0
    STOCK_TIMER_LB.place_forget()
    if stop_threads:
        STEXT.config(state="normal")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.insert(tk.END, "目前並無任務執行中.\n")
        STEXT.insert(tk.END, "=====================================\n")
        STEXT.see(tk.END)
        STEXT.config(state="disable")
    stop_threads = True
    PROGRESS["value"] = 100
    print("kill successful")

def align_center(root: Any, width: int, height: int) -> None:
    """ GUI視窗設定 """
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    size = '%dx%d+%d+%d' % (width, height, (screenwidth - width)/2, (screenheight - height)/2)
    root.geometry(size)

def f1_f1_selectall_event() -> None:
    """ 損益表，一般類別全選點擊事件 """
    for i in range(len(F1_1_CHECK_BTN_LIST)):
        globals()["f1_in_f1_btn"+str(i)].select()

def f1_f1_dselectall_event() -> None:
    """ 損益表，一般類別取消全選點擊事件 """
    for i in range(len(F1_1_CHECK_BTN_LIST)):
        globals()["f1_in_f1_btn"+str(i)].deselect()

def f1_f2_selectall_event() -> None:
    """ 損益表，保險類別全選點擊事件 """
    for i in range(len(F1_2_CHECK_BTN_LIST)):
        globals()["f1_in_f2_btn"+str(i)].select()

def f1_f2_dselectall_event() -> None:
    """ 損益表，保險類別取消全選點擊事件 """
    for i in range(len(F1_2_CHECK_BTN_LIST)):
        globals()["f1_in_f2_btn"+str(i)].deselect()

def f1_f3_selectall_event() -> None:
    """ 損益表，銀行類別全選點擊事件 """
    for i in range(len(F1_3_CHECK_BTN_LIST)):
        globals()["f1_in_f3_btn"+str(i)].select()

def f1_f3_dselectall_event() -> None:
    """ 損益表，銀行類別取消全選點擊事件 """
    for i in range(len(F1_3_CHECK_BTN_LIST)):
        globals()["f1_in_f3_btn"+str(i)].deselect()

def f2_f1_selectall_event() -> None:
    """ 資產負債表，一般類別全選點擊事件 """
    for i in range(len(F2_1_CHECK_BTN_LIST)):
        globals()["f2_in_f1_btn"+str(i)].select()

def f2_f1_dselectall_event() -> None:
    """ 資產負債表，一般類別取消全選點擊事件 """
    for i in range(len(F2_1_CHECK_BTN_LIST)):
        globals()["f2_in_f1_btn"+str(i)].deselect()

def f2_f2_selectall_event() -> None:
    """ 資產負債表，保險類別全選點擊事件 """
    for i in range(len(F2_2_CHECK_BTN_LIST)):
        globals()["f2_in_f2_btn"+str(i)].select()

def f2_f2_dselectall_event() -> None:
    """ 資產負債表，保險類別取消全選點擊事件 """
    for i in range(len(F2_2_CHECK_BTN_LIST)):
        globals()["f2_in_f2_btn"+str(i)].deselect()

def f2_f3_selectall_event() -> None:
    """ 資產負債表，銀行類別全選點擊事件 """
    for i in range(len(F2_3_CHECK_BTN_LIST)):
        globals()["f2_in_f3_btn"+str(i)].select()

def f2_f3_dselectall_event() -> None:
    """ 資產負債表，銀行類別取消全選點擊事件 """
    for i in range(len(F2_3_CHECK_BTN_LIST)):
        globals()["f2_in_f3_btn"+str(i)].deselect()

def f3_f1_selectall_event() -> None:
    """ 現金流量表，一般類別全選點擊事件 """
    for i in range(len(F3_1_CHECK_BTN_LIST)):
        globals()["f3_in_f1_btn"+str(i)].select()

def f3_f1_dselectall_event() -> None:
    """ 現金流量表，一般類別取消全選點擊事件 """
    for i in range(len(F3_1_CHECK_BTN_LIST)):
        globals()["f3_in_f1_btn"+str(i)].deselect()

def singledownload_path_event() -> None:
    """ 個股下載路徑點擊事件 """
    filename = filedialog.askdirectory(parent=window)
    if not filename:
        filename = "C:/"
    DOWNLOAD_PATH_TEXT.config(state="normal")
    DOWNLOAD_PATH_TEXT.delete(1.0,"end")
    DOWNLOAD_PATH_TEXT.insert(1.0, filename)
    DOWNLOAD_PATH_TEXT.config(state="disable")
    write_pathconfig(type="single",download_path=filename)

def bulkdownload_path_event() -> None:
    """ 批量下載路徑點擊事件 """
    filename = filedialog.askdirectory(parent=window)
    if not filename:
        filename = "C:/"
    BULK_DOWNLOAD_PATH_TEXT.config(state="normal")
    BULK_DOWNLOAD_PATH_TEXT.delete(1.0,"end")
    BULK_DOWNLOAD_PATH_TEXT.insert(1.0, filename)
    BULK_DOWNLOAD_PATH_TEXT.config(state="disable")
    write_pathconfig(type="bulk",download_path=filename)

def stocklistpath_event() -> None:
    """ 個股清單路徑點擊事件 """
    filename = filedialog.askopenfilename(parent=window,initialdir="C:/",filetypes = (("excel files","*.xlsx"),("all files","*.*")))
    STOCK_LIST_PATH_TEXT.config(state="normal")
    STOCK_LIST_PATH_TEXT.delete(1.0,"end")
    STOCK_LIST_PATH_TEXT.insert(1.0, filename)
    STOCK_LIST_PATH_TEXT.config(state="disable")

def f1_name_initialization() -> None:
    """ 損益表變數實例化 """
    for i in range(len(F1_1_VAR_NAME_lIST)):
        globals()["f1_in_f1_va"+str(i)] = tk.StringVar()
    for i in range(len(F1_2_VAR_NAME_lIST)):
        globals()["f1_in_f2_va"+str(i)] = tk.StringVar()
    for i in range(len(F1_3_VAR_NAME_lIST)):
        globals()["f1_in_f3_va"+str(i)] = tk.StringVar()

def f2_name_initialization() -> None:
    """ 資產負債表變數實例化 """
    for i in range(len(F2_1_VAR_NAME_lIST)):
        globals()["f2_in_f1_va"+str(i)] = tk.StringVar()
    for i in range(len(F2_2_VAR_NAME_lIST)):
        globals()["f2_in_f2_va"+str(i)] = tk.StringVar()
    for i in range(len(F2_3_VAR_NAME_lIST)):
        globals()["f2_in_f3_va"+str(i)] = tk.StringVar()

def f3_name_initialization() -> None:
    """ 現金流量表變數實例化 """
    for i in range(len(F3_1_VAR_NAME_lIST)):
        globals()["f3_in_f1_va"+str(i)] = tk.StringVar()

def frame1_check_btn() -> tuple[Callable, Callable, Callable]:
    """ 損益表複選單按鈕實例化 """
    frame1_1_list = FRAME1_1_INSIDE_LIST.copy()
    frame1_2_list = FRAME1_2_INSIDE_LIST.copy()
    frame1_3_list = FRAME1_3_INSIDE_LIST.copy()
    opt1 = tk.IntVar()
    opt2 = tk.IntVar()
    opt3 = tk.IntVar()
    FRAME_1_INSIDE_1_SALL_RBTN = tk.Radiobutton(FRAME_1_INSIDE_1_TEXT, variable=opt1,value=1,text="全選",
                                               command=f1_f1_selectall_event,bg="white",width=11,anchor="w",
                                               font=("microsoft yahei",8,"bold"))
    FRAME_1_INSIDE_1_DALL_RBTN = tk.Radiobutton(FRAME_1_INSIDE_1_TEXT, variable=opt1,value=2,text="取消全選",
                                               command=f1_f1_dselectall_event,bg="white",width=10,anchor="w",
                                               font=("microsoft yahei",8,"bold"))
    FRAME_1_INSIDE_2_SALL_RBTN = tk.Radiobutton(FRAME_1_INSIDE_2_TEXT, var=opt2,value=1,text="全選",
                                               command=f1_f2_selectall_event,bg="white",width=11,anchor="w",
                                               font=("microsoft yahei",8,"bold"))
    FRAME_1_INSIDE_2_DALL_RBTN = tk.Radiobutton(FRAME_1_INSIDE_2_TEXT, var=opt2,value=2,text="取消全選",
                                               command=f1_f2_dselectall_event,bg="white",width=10,anchor="w",
                                               font=("microsoft yahei",8,"bold"))
    FRAME_1_INSIDE_3_SALL_RBTN = tk.Radiobutton(FRAME_1_INSIDE_3_TEXT, var=opt3,value=1,text="全選",
                                               command=f1_f3_selectall_event,bg="white",width=11,anchor="w",
                                               font=("microsoft yahei",8,"bold"))
    FRAME_1_INSIDE_3_DALL_RBTN = tk.Radiobutton(FRAME_1_INSIDE_3_TEXT, var=opt3,value=2,text="取消全選",
                                               command=f1_f3_dselectall_event,bg="white",width=10,anchor="w",
                                               font=("microsoft yahei",8,"bold"))
    FRAME_1_INSIDE_1_TEXT.window_create("end", window=FRAME_1_INSIDE_1_SALL_RBTN)
    FRAME_1_INSIDE_1_TEXT.window_create("end", window=FRAME_1_INSIDE_1_DALL_RBTN)
    FRAME_1_INSIDE_2_TEXT.window_create("end", window=FRAME_1_INSIDE_2_SALL_RBTN)
    FRAME_1_INSIDE_2_TEXT.window_create("end", window=FRAME_1_INSIDE_2_DALL_RBTN)
    FRAME_1_INSIDE_3_TEXT.window_create("end", window=FRAME_1_INSIDE_3_SALL_RBTN)
    FRAME_1_INSIDE_3_TEXT.window_create("end", window=FRAME_1_INSIDE_3_DALL_RBTN)
    for i in range(len(F1_1_CHECK_BTN_LIST)):
        if frame1_1_list[0] in F1_1_BLUE_BTN_LIST:
            globals()["f1_in_f1_btn"+str(i)] = tk.Checkbutton(FRAME_1_INSIDE_1_TEXT,text=frame1_1_list[0],
                                                              var=globals()["f1_in_f1_va"+str(i)],
                                                              width=100,bg="white",fg="blue",anchor="w",
                                                              onvalue=frame1_1_list[0],offvalue="",
                                                              font=("microsoft yahei",8,"bold"))
            FRAME_1_INSIDE_1_TEXT.window_create("end", window=globals()["f1_in_f1_btn"+str(i)])
        else:
            globals()["f1_in_f1_btn"+str(i)] = tk.Checkbutton(FRAME_1_INSIDE_1_TEXT,text=frame1_1_list[0],
                                                              var=globals()["f1_in_f1_va"+str(i)],
                                                              width=100,bg="white",anchor="w",
                                                              onvalue=frame1_1_list[0],offvalue="",
                                                              font=("microsoft yahei",8,"bold"))
            FRAME_1_INSIDE_1_TEXT.window_create("end", window=globals()["f1_in_f1_btn"+str(i)])
        if len(frame1_1_list) >1:
            FRAME_1_INSIDE_1_TEXT.insert("end", "\n")
        frame1_1_list.pop(0)
    for i in range(len(F1_2_CHECK_BTN_LIST)):
        if frame1_2_list[0] in F1_2_BLUE_BTN_LIST:
            globals()["f1_in_f2_btn"+str(i)] = tk.Checkbutton(FRAME_1_INSIDE_2_TEXT,text=frame1_2_list[0],
                                                              var=globals()["f1_in_f2_va"+str(i)],
                                                              width=100,bg="white",fg="blue",anchor="w",
                                                              onvalue=frame1_2_list[0],offvalue="",
                                                              font=("microsoft yahei",8,"bold"))
            FRAME_1_INSIDE_2_TEXT.window_create("end", window=globals()["f1_in_f2_btn"+str(i)])
        else:
            globals()["f1_in_f2_btn"+str(i)] = tk.Checkbutton(FRAME_1_INSIDE_2_TEXT,text=frame1_2_list[0],
                                                              var=globals()["f1_in_f2_va"+str(i)],
                                                              width=100,bg="white",anchor="w",
                                                              onvalue=frame1_2_list[0],offvalue="",
                                                              font=("microsoft yahei",8,"bold"))
            FRAME_1_INSIDE_2_TEXT.window_create("end", window=globals()["f1_in_f2_btn"+str(i)])
        if len(frame1_2_list) >1:
            FRAME_1_INSIDE_2_TEXT.insert("end", "\n")
        frame1_2_list.pop(0)
    for i in range(len(F1_3_CHECK_BTN_LIST)):
        if frame1_3_list[0] in F1_3_BLUE_BTN_LIST:
            globals()["f1_in_f3_btn"+str(i)] = tk.Checkbutton(FRAME_1_INSIDE_3_TEXT,text=frame1_3_list[0],
                                                              var=globals()["f1_in_f3_va"+str(i)],
                                                              width=100,bg="white",fg="blue",anchor="w",
                                                              onvalue=frame1_3_list[0],offvalue="",
                                                              font=("microsoft yahei",8,"bold"))
            FRAME_1_INSIDE_3_TEXT.window_create("end", window=globals()["f1_in_f3_btn"+str(i)])
        else:
            globals()["f1_in_f3_btn"+str(i)] = tk.Checkbutton(FRAME_1_INSIDE_3_TEXT,text=frame1_3_list[0],
                                                              var=globals()["f1_in_f3_va"+str(i)],
                                                              width=100,bg="white",anchor="w",
                                                              onvalue=frame1_3_list[0],offvalue="",
                                                              font=("microsoft yahei",8,"bold"))
            FRAME_1_INSIDE_3_TEXT.window_create("end", window=globals()["f1_in_f3_btn"+str(i)])
        if len(frame1_3_list) >1:
            FRAME_1_INSIDE_3_TEXT.insert("end", "\n")
        frame1_3_list.pop(0)
    return FRAME_1_INSIDE_1_SALL_RBTN,FRAME_1_INSIDE_2_SALL_RBTN,FRAME_1_INSIDE_3_SALL_RBTN

def frame2_check_btn() -> tuple[Callable, Callable, Callable]:
    """ 資產負債表複選單按鈕實例化 """
    opt1 = tk.IntVar()
    opt2 = tk.IntVar()
    opt3 = tk.IntVar()
    FRAME_2_INSIDE_1_SALL_RBTN = tk.Radiobutton(FRAME_2_INSIDE_1_TEXT, variable=opt1, value=1, text="全選",
                                               command=f2_f1_selectall_event, bg="white", width=11, anchor="w",
                                               font=("microsoft yahei", 8, "bold"))
    FRAME_2_INSIDE_1_DALL_RBTN = tk.Radiobutton(FRAME_2_INSIDE_1_TEXT, variable=opt1, value=2, text="取消全選",
                                               command=f2_f1_dselectall_event, bg="white", width=10, anchor="w",
                                               font=("microsoft yahei", 8, "bold"))
    FRAME_2_INSIDE_2_SALL_RBTN = tk.Radiobutton(FRAME_2_INSIDE_2_TEXT, var=opt2, value=1, text="全選",
                                               command=f2_f2_selectall_event, bg="white", width=11, anchor="w",
                                               font=("microsoft yahei", 8, "bold"))
    FRAME_2_INSIDE_2_DALL_RBTN = tk.Radiobutton(FRAME_2_INSIDE_2_TEXT, var=opt2, value=2, text="取消全選",
                                               command=f2_f2_dselectall_event, bg="white", width=10, anchor="w",
                                               font=("microsoft yahei", 8, "bold"))
    FRAME_2_INSIDE_3_SALL_RBTN = tk.Radiobutton(FRAME_2_INSIDE_3_TEXT, var=opt3, value=1, text="全選",
                                               command=f2_f3_selectall_event, bg="white", width=11, anchor="w",
                                               font=("microsoft yahei", 8, "bold"))
    FRAME_2_INSIDE_3_DALL_RBTN = tk.Radiobutton(FRAME_2_INSIDE_3_TEXT, var=opt3, value=2, text="取消全選",
                                               command=f2_f3_dselectall_event, bg="white", width=10, anchor="w",
                                               font=("microsoft yahei", 8, "bold"))
    FRAME_2_INSIDE_1_TEXT.window_create("end", window=FRAME_2_INSIDE_1_SALL_RBTN)
    FRAME_2_INSIDE_1_TEXT.window_create("end", window=FRAME_2_INSIDE_1_DALL_RBTN)
    FRAME_2_INSIDE_2_TEXT.window_create("end", window=FRAME_2_INSIDE_2_SALL_RBTN)
    FRAME_2_INSIDE_2_TEXT.window_create("end", window=FRAME_2_INSIDE_2_DALL_RBTN)
    FRAME_2_INSIDE_3_TEXT.window_create("end", window=FRAME_2_INSIDE_3_SALL_RBTN)
    FRAME_2_INSIDE_3_TEXT.window_create("end", window=FRAME_2_INSIDE_3_DALL_RBTN)
    for i in range(len(F2_1_CHECK_BTN_LIST)):
        if FRAME2_1_INSIDE_LIST[0] in F2_1_BLUE_BTN_LIST:
            globals()["f2_in_f1_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_1_TEXT,text=FRAME2_1_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f1_va"+str(i)],
                                                              width=100, bg="white", fg="blue", anchor="w",
                                                              onvalue=FRAME2_1_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_1_TEXT.window_create("end", window=globals()["f2_in_f1_btn"+str(i)])
        elif FRAME2_1_INSIDE_LIST[0] in F2_1_GREEN_BTN_LIST:
            globals()["f2_in_f1_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_1_TEXT,text=FRAME2_1_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f1_va"+str(i)],
                                                              width=100, bg="white", fg="green", anchor="w",
                                                              onvalue=FRAME2_1_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_1_TEXT.window_create("end", window=globals()["f2_in_f1_btn"+str(i)])
        else:
            globals()["f2_in_f1_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_1_TEXT, text=FRAME2_1_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f1_va"+str(i)],
                                                              width=100, bg="white", anchor="w",
                                                              onvalue=FRAME2_1_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_1_TEXT.window_create("end", window=globals()["f2_in_f1_btn"+str(i)])
        if len(FRAME2_1_INSIDE_LIST) >1:
            FRAME_2_INSIDE_1_TEXT.insert("end", "\n")
        FRAME2_1_INSIDE_LIST.pop(0)
    for i in range(len(F2_2_CHECK_BTN_LIST)):
        if FRAME2_2_INSIDE_LIST[0] in F2_2_BLUE_BTN_LIST:
            globals()["f2_in_f2_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_2_TEXT, text=FRAME2_2_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f2_va"+str(i)],
                                                              width=100, bg="white", fg="blue", anchor="w",
                                                              onvalue=FRAME2_2_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_2_TEXT.window_create("end", window=globals()["f2_in_f2_btn"+str(i)])
        elif FRAME2_2_INSIDE_LIST[0] in F2_2_GREEN_BTN_LIST:
            globals()["f2_in_f2_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_2_TEXT, text=FRAME2_2_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f2_va"+str(i)],
                                                              width=100, bg="white", fg="green", anchor="w",
                                                              onvalue=FRAME2_2_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_2_TEXT.window_create("end", window=globals()["f2_in_f2_btn"+str(i)])
        else:
            globals()["f2_in_f2_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_2_TEXT, text=FRAME2_2_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f2_va"+str(i)],
                                                              width=100, bg="white", anchor="w",
                                                              onvalue=FRAME2_2_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_2_TEXT.window_create("end", window=globals()["f2_in_f2_btn"+str(i)])
        if len(FRAME2_2_INSIDE_LIST) >1:
            FRAME_2_INSIDE_2_TEXT.insert("end", "\n")
        FRAME2_2_INSIDE_LIST.pop(0)
    for i in range(len(F2_3_CHECK_BTN_LIST)):
        if FRAME2_3_INSIDE_LIST[0] in F2_3_BLUE_BTN_LIST:
            globals()["f2_in_f3_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_3_TEXT,text=FRAME2_3_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f3_va"+str(i)],
                                                              width=100, bg="white", fg="blue", anchor="w",
                                                              onvalue=FRAME2_3_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_3_TEXT.window_create("end", window=globals()["f2_in_f3_btn"+str(i)])
        elif FRAME2_3_INSIDE_LIST[0] in F2_3_GREEN_BTN_LIST:
            globals()["f2_in_f3_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_3_TEXT,text=FRAME2_3_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f3_va"+str(i)],
                                                              width=100, bg="white",fg="green", anchor="w",
                                                              onvalue=FRAME2_3_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_3_TEXT.window_create("end", window=globals()["f2_in_f3_btn"+str(i)])
        else:
            globals()["f2_in_f3_btn"+str(i)] = tk.Checkbutton(FRAME_2_INSIDE_3_TEXT,text=FRAME2_3_INSIDE_LIST[0],
                                                              var=globals()["f2_in_f3_va"+str(i)],
                                                              width=100, bg="white", anchor="w",
                                                              onvalue=FRAME2_3_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_2_INSIDE_3_TEXT.window_create("end", window=globals()["f2_in_f3_btn"+str(i)])
        if len(FRAME2_3_INSIDE_LIST) >1:
            FRAME_2_INSIDE_3_TEXT.insert("end", "\n")
        FRAME2_3_INSIDE_LIST.pop(0)
    return FRAME_2_INSIDE_1_SALL_RBTN,FRAME_2_INSIDE_2_SALL_RBTN,FRAME_2_INSIDE_3_SALL_RBTN

def frame3_check_btn() -> tuple[Callable]:
    """ 現金流量表複選單按鈕實例化 """
    opt1 = tk.IntVar()
    FRAME_3_INSIDE_1_SALL_RBTN = tk.Radiobutton(FRAME_3_INSIDE_1_TEXT, variable=opt1,
                                                value=1,text="全選", command=f3_f1_selectall_event,
                                                bg="white", width=11, anchor="w",
                                                font=("microsoft yahei", 8, "bold"))
    FRAME_3_INSIDE_1_DALL_RBTN = tk.Radiobutton(FRAME_3_INSIDE_1_TEXT, variable=opt1, value=2, text="取消全選",
                                               command=f3_f1_dselectall_event, bg="white", width=10, anchor="w",
                                               font=("microsoft yahei", 8, "bold"))
    FRAME_3_INSIDE_1_TEXT.window_create("end", window=FRAME_3_INSIDE_1_SALL_RBTN)
    FRAME_3_INSIDE_1_TEXT.window_create("end", window=FRAME_3_INSIDE_1_DALL_RBTN)
    for i in range(len(F3_1_CHECK_BTN_LIST)):
        if FRAME3_1_INSIDE_LIST[0] in F3_1_BLUE_BTN_LIST:
            globals()["f3_in_f1_btn"+str(i)] = tk.Checkbutton(FRAME_3_INSIDE_1_TEXT, text=FRAME3_1_INSIDE_LIST[0],
                                                              var=globals()["f3_in_f1_va"+str(i)],
                                                              width=100, bg="white", fg="blue", anchor="w",
                                                              onvalue=FRAME3_1_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_3_INSIDE_1_TEXT.window_create("end", window=globals()["f3_in_f1_btn"+str(i)])

        else:
            globals()["f3_in_f1_btn"+str(i)] = tk.Checkbutton(FRAME_3_INSIDE_1_TEXT,text=FRAME3_1_INSIDE_LIST[0],
                                                              var=globals()["f3_in_f1_va"+str(i)],
                                                              width=100, bg="white", anchor="w",
                                                              onvalue=FRAME3_1_INSIDE_LIST[0], offvalue="",
                                                              font=("microsoft yahei", 8, "bold"))
            FRAME_3_INSIDE_1_TEXT.window_create("end", window=globals()["f3_in_f1_btn"+str(i)])
        if len(FRAME3_1_INSIDE_LIST) >1:
            FRAME_3_INSIDE_1_TEXT.insert("end", "\n")
        FRAME3_1_INSIDE_LIST.pop(0)
    return FRAME_3_INSIDE_1_SALL_RBTN

align_center(window, 1000, 700)
window.update_idletasks()

"""GUI物件實例化"""


COUNTRY_LB = tk.Label(window, text="股票代號/國家： ",font=("新細明體", 12), anchor="e")
COUNTRY_LB.place(relx=0.01, rely=0.05, relwidth=0.15)

COUNTRY_BTN = ttk.Combobox(window,value=country_list)
COUNTRY_BTN.place(relx=0.16,rely=0.05,relwidth=0.3)
COUNTRY_BTN.current(0)

MARKET_LB = tk.Label(window, text="市場交易所： ", font=("新細明體", 12), anchor="e")
MARKET_LB.place(relx=0.01, rely=0.12, relwidth=0.15)

MARKET_BTN = ttk.Combobox(window, state="readonly")
MARKET_BTN.place(relx=0.16, rely=0.12, relwidth=0.3)

CATEGORY_LB = tk.Label(window,text="股票類別： ", font=("新細明體", 12), anchor="e")
CATEGORY_LB.place(relx=0.01, rely=0.19, relwidth=0.15)

CATEGORY_BTN = ttk.Combobox(window,state= "readonly")
CATEGORY_BTN.place(relx=0.16, rely=0.19, relwidth=0.3)

START_DATE_LB = tk.Label(window, text="開始日期： ", font=("新細明體", 12), anchor="e")
START_DATE_LB.place(relx=0.01, rely=0.26, relwidth=0.15)

YEAR, MONTH, DAY = get_dateconfig()
START_DATE_BTN = DateEntry(window,width=10,background='darkblue', year=YEAR, month=MONTH, day=DAY,
                           foreground="white", borderwidth=2, locale = "en_us", date_pattern ="yyyy/mm/dd")
START_DATE_BTN.place(relx=0.16,rely=0.26,relwidth=0.3)

END_DATE_LB = tk.Label(window, text="結束日期： ", font=("新細明體",12), anchor="e")
END_DATE_LB.place(relx=0.01, rely=0.33, relwidth=0.15)

END_DATE_BTN = DateEntry(window, width=10, background="darkblue", foreground="white", borderwidth=2, locale = "en_us", date_pattern ="yyyy/mm/dd")
END_DATE_BTN.place(relx=0.16, rely=0.33, relwidth=0.3)

STEXT = ScrolledText(window, bg="white", selectbackground="blue")
STEXT.config(state="disabled", font=("新細明體", 13))
STEXT.place(relx=0.01, rely=0.4, relwidth=0.49, relheight=0.2)

WINDOW_SPEA = ttk.Separator(window,orient="horizontal")
WINDOW_SPEA.place(relx=0, rely=0.62, relwidth=1)

DOWNLOAD_PATH_LB = tk.Label(window, text="個股下載路徑： ", font=("新細明體", 12), anchor="e")
DOWNLOAD_PATH_LB.place(relx=0.01, rely=0.632, relwidth=0.15)

DOWNLOAD_PATH_BTN = ttk.Button(window, text="選擇路徑", command=singledownload_path_event)
DOWNLOAD_PATH_BTN.place(relx=0.16, rely=0.63)

DOWNLOAD_PATH_TEXT = tk.Text(window)
DOWNLOAD_PATH_TEXT.config(state="normal")
DOWNLOAD_PATH_TEXT.insert(1.0, "C:/")
DOWNLOAD_PATH_TEXT.config(state="disabled")
DOWNLOAD_PATH_TEXT.place(relx=0.03, rely=0.7, relwidth=0.4, relheight=0.04)

BULK_DOWNLOAD_PATH_LB = tk.Label(window, text="批量下載路徑： ", font=("新細明體", 12), anchor="e")
BULK_DOWNLOAD_PATH_LB.place(relx=0.01, rely=0.755, relwidth=0.15)

BULK_DOWNLOAD_PATH_BTN = ttk.Button(window, text="選擇路徑", command=bulkdownload_path_event)
BULK_DOWNLOAD_PATH_BTN.place(relx=0.16, rely=0.75)

BULK_DOWNLOAD_PATH_TEXT = tk.Text(window)
BULK_DOWNLOAD_PATH_TEXT.config(state="normal")
BULK_DOWNLOAD_PATH_TEXT.insert(1.0, "C:/")
BULK_DOWNLOAD_PATH_TEXT.config(state="disabled")
BULK_DOWNLOAD_PATH_TEXT.place(relx=0.03, rely=0.82, relwidth=0.4, relheight=0.04)

STOCK_LIST_PATH_LB = tk.Label(window, text="指定清單檔案： ", font=("新細明體", 12), anchor="e")
STOCK_LIST_PATH_LB.place(relx=0.51, rely=0.635, relwidth=0.15)

STOCK_LIST_PATH_BTN = ttk.Button(window,text="選擇檔案", command=stocklistpath_event)
STOCK_LIST_PATH_BTN.place(relx=0.66, rely=0.63)

STOCK_LIST_PATH_TEXT = tk.Text(window)
STOCK_LIST_PATH_TEXT.config(state="disabled")
STOCK_LIST_PATH_TEXT.place(relx=0.53, rely=0.7, relwidth=0.4, relheight=0.04)

CRAWLER_STOCK_LIST_BTN = ttk.Button(window, text="下載個股清單", command=crawlerstocklist_btn_event)
CRAWLER_STOCK_LIST_BTN.place(relx=0.53, rely=0.8)

STOCK_TIMER_LB = tk.Label(window,text="", font=("新細明體", 10), anchor="e")
STOCK_TIMER_LB.place(relx=0.53, rely=0.88)

BULK_CRAWLER_BTN = ttk.Button(window, text="下載清單資料", command=bulkcrawler_btn_event)
BULK_CRAWLER_BTN.place(relx=0.68, rely=0.8)

BULK_START_BTN = ttk.Button(window, text="暫停下載", command=stopcrawler_btn_event)
BULK_START_BTN.place(relx=0.68, rely=0.87)

SINGLE_CRAWLER_BTN = ttk.Button(window,text="下載目標個股", command=singlecrawler_btn_event)
SINGLE_CRAWLER_BTN.place(relx=0.83, rely=0.8)

PROGRESS = ttk.Progressbar(window, style="red.Horizontal.TProgressbar", orient="horizontal", mode="determinate")
PROGRESS.place(relx=0.53 ,rely=0.94, relwidth=0.43)

""" notebook&frame實例化 """

NOTEBOOK = ttk.Notebook(window)
NOTEBOOK.place(relx=0.5, rely=0, relwidth=0.5, relheight=0.6)
FRAME_1 = tk.Frame(NOTEBOOK)
FRAME_2 = tk.Frame(NOTEBOOK)
FRAME_3 = tk.Frame(NOTEBOOK)

""" frame grid設定 """

NOTEBOOK.add(FRAME_1, text="損益表", padding=10)
NOTEBOOK.add(FRAME_2, text="資產負債表", padding=10)
NOTEBOOK.add(FRAME_3, text="現金流", padding=10)

""" frame1實例化 """

FRAME_1_NOTEBOOK = ttk.Notebook(FRAME_1)
FRAME_1_NOTEBOOK.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_1_INSIDE_1 = tk.Frame(FRAME_1_NOTEBOOK)
FRAME_1_INSIDE_1.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_1_INSIDE_1_SCOLLBAR = tk.Scrollbar(FRAME_1_INSIDE_1)
FRAME_1_INSIDE_1_SCOLLBAR.place(relx=0.95, relwidth=0.05, relheight=1)
FRAME_1_INSIDE_1_TEXT = tk.Text(FRAME_1_INSIDE_1)
FRAME_1_INSIDE_1_TEXT.place(relx=0, rely=0, relwidth=0.95, relheight=1)
FRAME_1_INSIDE_1_TEXT.configure(state="disabled")
FRAME_1_INSIDE_1_SCOLLBAR.config(command=FRAME_1_INSIDE_1_TEXT.yview)
FRAME_1_INSIDE_1_TEXT.config(yscrollcommand=FRAME_1_INSIDE_1_SCOLLBAR.set)
FRAME_1_INSIDE_2 = tk.Frame(FRAME_1_NOTEBOOK)
FRAME_1_INSIDE_2.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_1_INSIDE_2_SCOLLBAR = tk.Scrollbar(FRAME_1_INSIDE_2)
FRAME_1_INSIDE_2_SCOLLBAR.place(relx=0.95, relwidth=0.05, relheight=1)
FRAME_1_INSIDE_2_TEXT = tk.Text(FRAME_1_INSIDE_2)
FRAME_1_INSIDE_2_TEXT.place(relx=0, rely=0, relwidth=0.95, relheight=1)
FRAME_1_INSIDE_2_TEXT.configure(state="disabled")
FRAME_1_INSIDE_2_SCOLLBAR.config(command=FRAME_1_INSIDE_2_TEXT.yview)
FRAME_1_INSIDE_2_TEXT.config(yscrollcommand=FRAME_1_INSIDE_2_SCOLLBAR.set)
FRAME_1_INSIDE_3 = tk.Frame(FRAME_1_NOTEBOOK)
FRAME_1_INSIDE_3.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_1_INSIDE_3_SCOLLBAR = tk.Scrollbar(FRAME_1_INSIDE_3)
FRAME_1_INSIDE_3_SCOLLBAR.place(relx=0.95, relwidth=0.05, relheight=1)
FRAME_1_INSIDE_3_TEXT = tk.Text(FRAME_1_INSIDE_3)
FRAME_1_INSIDE_3_TEXT.place(relx=0, rely=0, relwidth=0.95, relheight=1)
FRAME_1_INSIDE_3_TEXT.configure(state="disabled")
FRAME_1_INSIDE_3_SCOLLBAR.config(command=FRAME_1_INSIDE_3_TEXT.yview)
FRAME_1_INSIDE_3_TEXT.config(yscrollcommand=FRAME_1_INSIDE_3_SCOLLBAR.set)
f1_name_initialization()
FRAME_1_INSIDE_1_SALL_RBTN, FRAME_1_INSIDE_2_SALL_RBTN, FRAME_1_INSIDE_3_SALL_RBTN = frame1_check_btn()
FRAME_1_NOTEBOOK.add(FRAME_1_INSIDE_1, text="一般類別")
FRAME_1_NOTEBOOK.add(FRAME_1_INSIDE_2, text="保險類別")
FRAME_1_NOTEBOOK.add(FRAME_1_INSIDE_3, text="銀行類別")

""" frame2實例化 """

FRAME_2_NOTEBOOK = ttk.Notebook(FRAME_2)
FRAME_2_NOTEBOOK.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_2_INSIDE_1 = tk.Frame(FRAME_2_NOTEBOOK)
FRAME_2_INSIDE_1.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_2_INSIDE_1_SCOLLBAR = tk.Scrollbar(FRAME_2_INSIDE_1)
FRAME_2_INSIDE_1_SCOLLBAR.place(relx=0.95, relwidth=0.05, relheight=1)
FRAME_2_INSIDE_1_TEXT = tk.Text(FRAME_2_INSIDE_1)
FRAME_2_INSIDE_1_TEXT.place(relx=0 ,rely=0, relwidth=0.95, relheight=1)
FRAME_2_INSIDE_1_TEXT.configure(state="disabled")
FRAME_2_INSIDE_1_SCOLLBAR.config(command=FRAME_2_INSIDE_1_TEXT.yview)
FRAME_2_INSIDE_1_TEXT.config(yscrollcommand=FRAME_2_INSIDE_1_SCOLLBAR.set)
FRAME_2_INSIDE_2 = tk.Frame(FRAME_2_NOTEBOOK)
FRAME_2_INSIDE_2.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_2_INSIDE_2_SCOLLBAR = tk.Scrollbar(FRAME_2_INSIDE_2)
FRAME_2_INSIDE_2_SCOLLBAR.place(relx=0.95, relwidth=0.05, relheight=1)
FRAME_2_INSIDE_2_TEXT = tk.Text(FRAME_2_INSIDE_2)
FRAME_2_INSIDE_2_TEXT.place(relx=0, rely=0, relwidth=0.95, relheight=1)
FRAME_2_INSIDE_2_TEXT.configure(state="disabled")
FRAME_2_INSIDE_2_SCOLLBAR.config(command=FRAME_2_INSIDE_2_TEXT.yview)
FRAME_2_INSIDE_2_TEXT.config(yscrollcommand=FRAME_2_INSIDE_2_SCOLLBAR.set)
FRAME2_INSIDE_3 = tk.Frame(FRAME_2_NOTEBOOK)
FRAME2_INSIDE_3.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_2_INSIDE_3_SCOLLBAR = tk.Scrollbar(FRAME2_INSIDE_3)
FRAME_2_INSIDE_3_SCOLLBAR.place(relx=0.95, relwidth=0.05, relheight=1)
FRAME_2_INSIDE_3_TEXT = tk.Text(FRAME2_INSIDE_3)
FRAME_2_INSIDE_3_TEXT.place(relx=0, rely=0, relwidth=0.95, relheight=1)
FRAME_2_INSIDE_3_TEXT.configure(state="disabled")
FRAME_2_INSIDE_3_SCOLLBAR.config(command=FRAME_2_INSIDE_3_TEXT.yview)
FRAME_2_INSIDE_3_TEXT.config(yscrollcommand=FRAME_2_INSIDE_3_SCOLLBAR.set)
f2_name_initialization()
FRAME_2_INSIDE_1_SALL_RBTN, FRAME_2_INSIDE_2_SALL_RBTN, FRAME_2_INSIDE_3_SALL_RBTN = frame2_check_btn()
FRAME_2_NOTEBOOK.add(FRAME_2_INSIDE_1, text="一般類別")
FRAME_2_NOTEBOOK.add(FRAME_2_INSIDE_2, text="保險類別")
FRAME_2_NOTEBOOK.add(FRAME2_INSIDE_3, text="銀行類別")

""" frame3實例化 """

FRAME_3_NOTEBOOK = ttk.Notebook(FRAME_3)
FRAME_3_NOTEBOOK.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_3_INSIDE_1 = tk.Frame(FRAME_3_NOTEBOOK)
FRAME_3_INSIDE_1.place(relx=0, rely=0, relwidth=1, relheight=1)
FRAME_3_INSIDE_1_SCOLLBAR = tk.Scrollbar(FRAME_3_INSIDE_1)
FRAME_3_INSIDE_1_SCOLLBAR.place(relx=0.95, relwidth=0.05, relheight=1)
FRAME_3_INSIDE_1_TEXT = tk.Text(FRAME_3_INSIDE_1)
FRAME_3_INSIDE_1_TEXT.place(relx=0, rely=0, relwidth=0.95, relheight=1)
FRAME_3_INSIDE_1_TEXT.configure(state="disabled")
FRAME_3_INSIDE_1_SCOLLBAR.config(command=FRAME_3_INSIDE_1_TEXT.yview)
FRAME_3_INSIDE_1_TEXT.config(yscrollcommand=FRAME_3_INSIDE_1_SCOLLBAR.set)
f3_name_initialization()
FRAME_3_INSIDE_1_SALL_RBTN = frame3_check_btn()
FRAME_3_NOTEBOOK.add(FRAME_3_INSIDE_1, text="一般類別")

""" 事件綁定 """

START_DATE_BTN.bind("<<DateEntrySelected>>", start_date_btn_event)
COUNTRY_BTN.bind("<<ComboboxSelected>>", country_btn_event)
COUNTRY_BTN.bind("<Return>", country_btn_enter_event)

"""複選單按鈕實例化"""

f1_1config, f1_2config, f1_3config, f2_1config, f2_2config, f2_3config, f3config = get_checkboxconfig()

for i in range(len(F1_1_CHECK_BTN_LIST)):
    if f1_1config[i]:
        globals()["f1_in_f1_btn"+str(i)].select()
    else:
        globals()["f1_in_f1_btn"+str(i)].deselect()
for i in range(len(F1_2_CHECK_BTN_LIST)):
    if f1_2config[i]:
        globals()["f1_in_f2_btn"+str(i)].select()
    else:
        globals()["f1_in_f2_btn"+str(i)].deselect()
for i in range(len(F1_3_CHECK_BTN_LIST)):
    if f1_3config[i]:
        globals()["f1_in_f3_btn"+str(i)].select()
    else:
        globals()["f1_in_f3_btn"+str(i)].deselect()
for i in range(len(F2_1_CHECK_BTN_LIST)):
    if f2_1config[i]:
        globals()["f2_in_f1_btn"+str(i)].select()
    else:
        globals()["f2_in_f1_btn"+str(i)].deselect()
for i in range(len(F2_2_CHECK_BTN_LIST)):
    if f2_2config[i]:
        globals()["f2_in_f2_btn"+str(i)].select()
    else:
        globals()["f2_in_f2_btn"+str(i)].deselect()
for i in range(len(F2_3_CHECK_BTN_LIST)):
    if f2_3config[i]:
        globals()["f2_in_f3_btn"+str(i)].select()
    else:
        globals()["f2_in_f3_btn"+str(i)].deselect()
for i in range(len(F3_1_CHECK_BTN_LIST)):
    if f3config[i]:
        globals()["f3_in_f1_btn"+str(i)].select()
    else:
        globals()["f3_in_f1_btn"+str(i)].deselect()

"""複選單按鈕預設全選"""

FRAME_1_INSIDE_1_SALL_RBTN.select()
FRAME_1_INSIDE_2_SALL_RBTN.select()
FRAME_1_INSIDE_3_SALL_RBTN.select()
FRAME_2_INSIDE_1_SALL_RBTN.select()
FRAME_2_INSIDE_2_SALL_RBTN.select()
FRAME_2_INSIDE_3_SALL_RBTN.select()
FRAME_3_INSIDE_1_SALL_RBTN.select()

""" 讀取存檔的變數設定 """

single,bulk = get_pathconfig()
DOWNLOAD_PATH_TEXT.config(state="normal")
DOWNLOAD_PATH_TEXT.delete(1.0, "end")
DOWNLOAD_PATH_TEXT.insert(1.0, single)
DOWNLOAD_PATH_TEXT.config(state="disable")
BULK_DOWNLOAD_PATH_TEXT.config(state="normal")
BULK_DOWNLOAD_PATH_TEXT.delete(1.0, "end")
BULK_DOWNLOAD_PATH_TEXT.insert(1.0, bulk)
BULK_DOWNLOAD_PATH_TEXT.config(state="disable")

if __name__ == "__main__":
    window.mainloop()
